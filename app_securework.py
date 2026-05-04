"""
SecureWork — Servidor Web Flask
Conecta directamente con la lógica del escritorio (codigos_completos_mod.py).
Ejecutar: python app.py
Abrir:    http://localhost:5000
"""

import cv2
import mediapipe as mp
import numpy as np
import time
import csv
import os
import sys
import json
import queue
import base64
import threading
import tempfile
import io
from datetime import datetime
from collections import defaultdict
from flask import (Flask, request, jsonify,
                   Response, send_file, session)
from functools import wraps

# ── Variables de modelos (se cargan en hilo de fondo) ────────
model_general         = None
model_epp             = None
face_mesh             = None
YOLO_OK               = False
EPP_MODELO_DISPONIBLE = False
modelos_listos        = False   # bandera que el frontend puede consultar

def _cargar_modelos():
    """Carga todos los modelos pesados en un hilo de fondo.
    Flask arranca inmediatamente y el login funciona al instante.
    Los modelos estarán listos unos segundos después."""
    global model_general, model_epp, face_mesh
    global YOLO_OK, EPP_MODELO_DISPONIBLE, modelos_listos

    print("[INFO] Cargando modelos en segundo plano...")

    # Mediapipe
    try:
        import mediapipe as mp
        _mp_face_mesh = mp.solutions.face_mesh
        face_mesh     = _mp_face_mesh.FaceMesh(refine_landmarks=True)
        print("[INFO] MediaPipe OK")
    except Exception as e:
        print(f"[AVISO] MediaPipe no disponible: {e}")
        face_mesh = None

    # YOLO general
    try:
        from ultralytics import YOLO
        model_general = YOLO("yolov8n.pt")
        model_general.overrides['conf'] = 0.55
        model_general.overrides['iou']  = 0.45
        YOLO_OK = True
        print("[INFO] YOLO general OK")
    except Exception as e:
        print(f"[AVISO] YOLO no disponible: {e}")
        YOLO_OK = False
        model_general = None

    # YOLO EPP
    try:
        from ultralytics import YOLO
        from huggingface_hub import login, hf_hub_download
        login(token="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")
        ckpt_path = hf_hub_download(
            repo_id="Hansung-Cho/yolov8-ppe-detection", filename="best.pt")
        model_epp = YOLO(ckpt_path)
        model_epp.overrides['conf']         = 0.25
        model_epp.overrides['iou']          = 0.45
        model_epp.overrides['agnostic_nms'] = False
        model_epp.overrides['max_det']      = 1000
        EPP_MODELO_DISPONIBLE = True
        print("[INFO] Modelo EPP OK")
    except Exception as e:
        print(f"[AVISO] Modelo EPP no disponible: {e}")
        model_epp = model_general
        EPP_MODELO_DISPONIBLE = False

    modelos_listos = True
    print("[INFO] ✓ Todos los modelos cargados. Sistema listo.")

# Arrancar carga en hilo separado (no bloquea Flask)
threading.Thread(target=_cargar_modelos, daemon=True).start()

# ── Constantes ────────────────────────────────────────────────
LEFT_EYE      = [33, 160, 158, 133, 153, 144]
RIGHT_EYE     = [362, 385, 387, 263, 373, 380]
EAR_UMBRAL    = 0.25
TIEMPO_UMBRAL = 5.0
CSV_REGISTROS = "registros_operarios.csv"
CSV_EVENTOS   = "eventos.txt"
EPP_NEGATIVO  = {"no-hardhat": "CASCO", "no-safety vest": "CHALECO"}
EPP_REQUERIDO = {"CASCO", "CHALECO"}
EPP_TIEMPO_GRACIA = 4.0
CAM_OPERADOR_1 = 0
CAM_OPERADOR_2 = 1

USUARIOS = {
    "admin":    {"password": "admin123", "rol": "administrador"},
    "empleado": {"password": "emp123",   "rol": "empleado"},
}

# ── Estado global compartido ──────────────────────────────────
tiempo_inicio_ojos = None
alarma_fatiga      = False
alarma_telefono    = False

# Estado independiente para cámara 2
tiempo_inicio_ojos2 = None
alarma_fatiga2      = False
alarma_telefono2    = False

# Soportamos 2 operarios simultáneos
# operarios[0] → asignado a cámara 1, operarios[1] → asignado a cámara 2
operarios      = [None, None]   # cada uno: dict {cedula, nombre, cargo} o None
horas_entrada  = [None, None]

eventos_recientes  = []
stats_eventos      = defaultdict(int)
_lock_estado       = threading.Lock()

# Estado de cámara 1 (Operador 1 / Admin)
cam1 = {
    "running":   False,
    "cap":       None,
    "thread":    None,
    "frame_jpg": None,   # bytes JPEG del último frame
    "ojos":      "SIN DETECCION",
    "ear":       0.0,
    "persona":   False,
    "telefono":  False,
    "epp_ok":    set(),
    "epp_falta": set(EPP_REQUERIDO),
}

# Estado de cámara 2 (Operador 2 — con análisis completo igual que cam1)
cam2 = {
    "running":   False,
    "cap":       None,
    "thread":    None,
    "frame_jpg": None,
    "ojos":      "SIN DETECCION",
    "ear":       0.0,
    "persona":   False,
    "telefono":  False,
    "epp_ok":    set(),
    "epp_falta": set(EPP_REQUERIDO),
}

# ── Estabilizadores ───────────────────────────────────────────
class EstabilizadorDeteccion:
    def __init__(self, fa=8, fd=20):
        self._fa = fa; self._fd = fd
        self._estado = False; self._cnt = 0; self._cand = False
    def actualizar(self, det):
        if det == self._cand:
            self._cnt += 1
        else:
            self._cand = det; self._cnt = 1
        umbral = self._fa if self._cand else self._fd
        if self._cnt >= umbral:
            self._estado = self._cand
        return self._estado
    def reconf(self, fa, fd):
        self._fa = fa; self._fd = fd

est_persona  = EstabilizadorDeteccion(6,  20)
est_casco    = EstabilizadorDeteccion(8,  30)
est_chaleco  = EstabilizadorDeteccion(8,  30)
est_telefono = EstabilizadorDeteccion(5,  8)

# Estabilizadores independientes para cámara 2
est2_persona  = EstabilizadorDeteccion(6,  20)
est2_casco    = EstabilizadorDeteccion(8,  30)
est2_chaleco  = EstabilizadorDeteccion(8,  30)
est2_telefono = EstabilizadorDeteccion(5,  8)

# ── Helpers CSV ───────────────────────────────────────────────
def inicializar_csv():
    if not os.path.exists(CSV_REGISTROS):
        with open(CSV_REGISTROS, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(
                ["Cedula","Nombre","Cargo","Entrada","Salida","Duracion_min"])

def guardar_evento(tipo, cedula=None):
    global eventos_recientes
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if cedula is None:
        op = operarios[0] or operarios[1]
        cedula = op["cedula"] if op else "DESCONOCIDO"
    with open(CSV_EVENTOS, "a", encoding="utf-8") as f:
        f.write(f"{ts} | {cedula} | {tipo}\n")
    ev = {"ts": datetime.now().strftime("%H:%M:%S"), "cedula": cedula, "tipo": tipo}
    eventos_recientes.insert(0, ev)
    if len(eventos_recientes) > 100:
        eventos_recientes = eventos_recientes[:100]
    stats_eventos[tipo.split(":")[0]] += 1

# ── Grabación de evidencias en video ──────────────────────────
EVIDENCIAS_DIR = os.path.join(os.path.expanduser("~"), "Desktop",
                              "SecureWork_Descargas", "evidencias")
DURACION_CLIP_SEG = 12   # segundos que graba el clip
CLIPS_ACTIVOS = {}       # {(slot, tipo_evento): True}  para evitar clips duplicados
_lock_clips = threading.Lock()

def _guardar_evidencia_json(carpeta_emp, cedula, nombre, cargo, ts_str, archivo, tipo):
    """Agrega una entrada al índice JSON de evidencias del operario."""
    idx_path = os.path.join(carpeta_emp, "evidencias.json")
    entrada = {
        "cedula":  cedula,
        "nombre":  nombre,
        "cargo":   cargo,
        "ts":      ts_str,
        "tipo":    tipo,
        "archivo": archivo,
    }
    try:
        if os.path.exists(idx_path):
            with open(idx_path, "r", encoding="utf-8") as f:
                items = json.load(f)
        else:
            items = []
        items.insert(0, entrada)   # más reciente primero
        with open(idx_path, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[AVISO] No se pudo actualizar evidencias.json: {e}")

def _grabar_clip(slot, tipo_evento, frames_buffer):
    """
    Graba un clip AVI a partir de los frames del buffer más los que vengan después.
    Se ejecuta en un hilo independiente para no bloquear el loop principal.

    Args:
        slot          : 0 → cámara 1, 1 → cámara 2
        tipo_evento   : "FATIGA", "TELEFONO", "EPP_FALTANTE:CASCO", etc.
        frames_buffer : list[np.ndarray] — frames previos al evento (aprox. 3 s)
    """
    clave = (slot, tipo_evento.split(":")[0])
    with _lock_clips:
        if clave in CLIPS_ACTIVOS:
            return          # ya hay un clip activo para este slot+tipo
        CLIPS_ACTIVOS[clave] = True

    try:
        cam = cam1 if slot == 0 else cam2
        op  = operarios[slot]
        if op is None:
            op = {"cedula": "DESCONOCIDO", "nombre": "Desconocido", "cargo": "N/A"}

        cedula = op["cedula"]
        nombre = op["nombre"]
        cargo  = op.get("cargo", "N/A")

        # Nombre del archivo: cedula_TIPO_FECHA_HORA.avi
        ts_dt   = datetime.now()
        ts_str  = ts_dt.strftime("%Y-%m-%d %H:%M:%S")
        ts_file = ts_dt.strftime("%Y%m%d_%H%M%S")
        tipo_safe = tipo_evento.replace(":", "-").replace(" ", "_")
        nombre_safe = nombre.replace(" ", "_")[:20]
        archivo = f"{cedula}_{tipo_safe}_{ts_file}.avi"

        # Carpeta: EVIDENCIAS_DIR/<cedula>_<nombre>/
        carpeta_emp = os.path.join(EVIDENCIAS_DIR, f"{cedula}_{nombre_safe}")
        os.makedirs(carpeta_emp, exist_ok=True)
        ruta_video = os.path.join(carpeta_emp, archivo)

        # Resolver dimensiones del primer frame disponible
        h, w = 480, 640
        muestra = frames_buffer[0] if frames_buffer else None
        if muestra is None and cam.get("frame_jpg"):
            buf = np.frombuffer(cam["frame_jpg"], dtype=np.uint8)
            muestra = cv2.imdecode(buf, cv2.IMREAD_COLOR)
        if muestra is not None:
            h, w = muestra.shape[:2]

        fourcc = cv2.VideoWriter_fourcc(*"XVID")
        fps_out = 10
        writer  = cv2.VideoWriter(ruta_video, fourcc, fps_out, (w, h))

        # ── Escribir frames del buffer pre-evento ──────────────
        for fr in frames_buffer:
            if fr.shape[1] != w or fr.shape[0] != h:
                fr = cv2.resize(fr, (w, h))
            writer.write(fr)

        # ── Grabar durante DURACION_CLIP_SEG ──────────────────
        t_fin = time.time() + DURACION_CLIP_SEG
        while time.time() < t_fin and cam.get("running", False):
            jpg = cam.get("frame_jpg")
            if jpg:
                buf   = np.frombuffer(jpg, dtype=np.uint8)
                frame = cv2.imdecode(buf, cv2.IMREAD_COLOR)
                if frame is not None:
                    if frame.shape[1] != w or frame.shape[0] != h:
                        frame = cv2.resize(frame, (w, h))
                    # ── Overlay de información ─────────────────
                    overlay_ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    overlay_op    = f"Operario: {nombre} | Cedula: {cedula} | {cargo}"
                    overlay_cam   = f"Camara {slot+1}"
                    overlay_evento= f"EVENTO: {tipo_evento}"
                    # Fondo semitransparente superior
                    cv2.rectangle(frame, (0, 0), (w, 56), (0, 0, 0), -1)
                    cv2.putText(frame, overlay_op,    (8, 16),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.44, (255, 255, 255), 1)
                    cv2.putText(frame, overlay_ts + "  |  " + overlay_cam,
                                (8, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.44, (180, 180, 255), 1)
                    # Banner de evento en color (rojo / amarillo / naranja)
                    if "FATIGA" in tipo_evento:
                        color_ev = (0, 0, 220)
                    elif "TELEFONO" in tipo_evento:
                        color_ev = (220, 180, 0)
                    else:
                        color_ev = (0, 110, 255)
                    cv2.rectangle(frame, (0, 38), (w, 56), color_ev, -1)
                    cv2.putText(frame, overlay_evento, (8, 52),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.44, (255, 255, 255), 1)
                    writer.write(frame)
            time.sleep(1.0 / fps_out)

        writer.release()
        print(f"[EVIDENCIA] Clip guardado: {ruta_video}")

        # ── Registrar en el índice JSON ────────────────────────
        _guardar_evidencia_json(carpeta_emp, cedula, nombre, cargo,
                                ts_str, archivo, tipo_evento)

    except Exception as e:
        print(f"[ERROR] _grabar_clip slot={slot} tipo={tipo_evento}: {e}")
    finally:
        with _lock_clips:
            CLIPS_ACTIVOS.pop(clave, None)


def registrar_entrada_fn(datos, slot):
    """Registra entrada en el slot indicado (0 o 1)."""
    global operarios, horas_entrada
    operarios[slot]     = datos
    horas_entrada[slot] = datetime.now()
    guardar_evento("ENTRADA", cedula=datos["cedula"])

def registrar_salida_fn(cedula):
    """Registra salida buscando la cedula en cualquier slot activo. Retorna slot liberado."""
    global operarios, horas_entrada
    for slot in range(2):
        op = operarios[slot]
        if op and op["cedula"] == cedula:
            hora_salida = datetime.now()
            dur = round((hora_salida - horas_entrada[slot]).total_seconds() / 60, 1)
            with open(CSV_REGISTROS, "a", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow([
                    op["cedula"], op["nombre"], op["cargo"],
                    horas_entrada[slot].strftime("%Y-%m-%d %H:%M:%S"),
                    hora_salida.strftime("%Y-%m-%d %H:%M:%S"), dur])
            guardar_evento("SALIDA", cedula=op["cedula"])
            operarios[slot]     = None
            horas_entrada[slot] = None
            return slot
    return None

def cargar_registros():
    if not os.path.exists(CSV_REGISTROS):
        return []
    with open(CSV_REGISTROS, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def cargar_eventos():
    if not os.path.exists(CSV_EVENTOS):
        return []
    eventos = []
    with open(CSV_EVENTOS, "r", encoding="utf-8") as f:
        for line in f:
            p = line.strip().split(" | ")
            if len(p) == 3:
                eventos.append({"ts": p[0], "cedula": p[1], "tipo": p[2]})
    return list(reversed(eventos))

def calcular_EAR(eye_points, landmarks, w, h):
    coords = [(int(landmarks[p].x*w), int(landmarks[p].y*h)) for p in eye_points]
    A = np.linalg.norm(np.array(coords[1])-np.array(coords[5]))
    B = np.linalg.norm(np.array(coords[2])-np.array(coords[4]))
    C = np.linalg.norm(np.array(coords[0])-np.array(coords[3]))
    return (A+B)/(2.0*C) if C > 0 else 0

def alarma_sonido():
    try:
        if sys.platform == "win32":
            import winsound
            for i in range(5):
                winsound.Beep(800+i*200, 200); time.sleep(0.1)
    except Exception:
        pass

def _filtrar_por_periodo(eventos, registros, periodo,
                         fecha_desde=None, fecha_hasta=None):
    from datetime import timedelta
    ahora = datetime.now()

    # Rango de fecha exacta tiene prioridad sobre el período
    if fecha_desde or fecha_hasta:
        try:
            inicio = datetime.strptime(fecha_desde, "%Y-%m-%d") if fecha_desde \
                     else datetime(2000, 1, 1)
            # fecha_hasta incluye todo el día (hasta las 23:59:59)
            fin = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59) if fecha_hasta \
                else datetime(2100, 12, 31)
        except ValueError:
            return eventos, registros

        ev_f = [e for e in eventos
                if _ts_en_rango(e.get("ts", ""), inicio, fin)]
        re_f = [r for r in registros
                if _ts_en_rango(r.get("Entrada", ""), inicio, fin)]
        return ev_f, re_f

    if periodo == "Día":
        inicio = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "Semana":
        inicio = (ahora - timedelta(days=ahora.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "Mes":
        inicio = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif periodo == "Año":
        inicio = ahora.replace(month=1, day=1, hour=0, minute=0,
                                second=0, microsecond=0)
    else:
        return eventos, registros

    ev_f = [e for e in eventos if _ts_ok(e.get("ts",""), inicio)]
    re_f = [r for r in registros if _ts_ok(r.get("Entrada",""), inicio)]
    return ev_f, re_f


def _ts_en_rango(ts_str, inicio, fin):
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(ts_str, fmt)
            return inicio <= dt <= fin
        except ValueError:
            pass
    return False

def _ts_ok(ts_str, inicio):
    for fmt in ("%Y-%m-%d %H:%M:%S", "%H:%M:%S"):
        try:
            dt = datetime.strptime(ts_str, fmt)
            if fmt == "%H:%M:%S":
                dt = dt.replace(year=datetime.now().year,
                                month=datetime.now().month,
                                day=datetime.now().day)
            return dt >= inicio
        except ValueError:
            pass
    return False

# ── Loop cámara 1 (con análisis completo) ────────────────────
def _loop_cam1():
    global tiempo_inicio_ojos, alarma_fatiga, alarma_telefono
    alarmas_epp   = set()
    t_epp_falta   = {}

    # Buffer circular de frames crudos (últimos ~3 s a ~10 fps → 30 frames)
    from collections import deque
    _buf1 = deque(maxlen=30)

    while cam1["running"]:
        ret, frame = cam1["cap"].read()
        if not ret:
            break

        h, w, _ = frame.shape
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        _buf1.append(frame.copy())   # guardar frame crudo en buffer

        # Fatiga
        estado_ojos  = "SIN DETECCION"
        ear_val      = 0.0
        if face_mesh is not None:
            results_face = face_mesh.process(rgb)
            if results_face.multi_face_landmarks:
                for lm in results_face.multi_face_landmarks:
                    le = calcular_EAR(LEFT_EYE,  lm.landmark, w, h)
                    re = calcular_EAR(RIGHT_EYE, lm.landmark, w, h)
                    ear_val = (le + re) / 2.0
                    if ear_val < EAR_UMBRAL:
                        estado_ojos = "OJOS CERRADOS"
                        if tiempo_inicio_ojos is None:
                            tiempo_inicio_ojos = time.time()
                        if time.time() - tiempo_inicio_ojos >= TIEMPO_UMBRAL:
                            estado_ojos = "!! FATIGA DETECTADA !!"
                            if not alarma_fatiga:
                                threading.Thread(target=alarma_sonido, daemon=True).start()
                                guardar_evento("FATIGA")
                                alarma_fatiga = True
                                threading.Thread(
                                    target=_grabar_clip,
                                    args=(0, "FATIGA", list(_buf1)),
                                    daemon=True).start()
                    else:
                        estado_ojos = "ACTIVO"
                        tiempo_inicio_ojos = None
                        alarma_fatiga = False

        # YOLO general
        persona_cruda  = False
        telefono_crudo = False
        if YOLO_OK and model_general:
            for r in model_general(frame, verbose=False):
                for box in r.boxes:
                    cls = int(box.cls[0])
                    nm  = model_general.names[cls].lower()
                    if cls == 0: persona_cruda = True
                    if "cell phone" in nm or "phone" in nm: telefono_crudo = True
                    x1,y1,x2,y2 = map(int, box.xyxy[0])
                    cv2.rectangle(frame,(x1,y1),(x2,y2),(255,220,0),2)
                    cv2.putText(frame, model_general.names[cls],
                                (x1,y1-5), cv2.FONT_HERSHEY_SIMPLEX, 0.5,(255,220,0),1)

        persona_det  = est_persona.actualizar(persona_cruda)
        telefono_det = est_telefono.actualizar(telefono_crudo)
        if telefono_det and not alarma_telefono:
            threading.Thread(target=alarma_sonido, daemon=True).start()
            guardar_evento("TELEFONO")
            alarma_telefono = True
            threading.Thread(
                target=_grabar_clip,
                args=(0, "TELEFONO", list(_buf1)),
                daemon=True).start()
        elif not telefono_det:
            alarma_telefono = False

        # YOLO EPP
        casco_crudo = chaleco_crudo = False
        if YOLO_OK and model_epp:
            for r in model_epp(frame, verbose=False):
                for box in r.boxes:
                    cls = int(box.cls[0])
                    nm  = model_epp.names[cls].lower()
                    if nm == "hardhat":     casco_crudo   = True
                    if nm == "safety vest": chaleco_crudo = True
                    es_neg = any(k in nm for k in EPP_NEGATIVO)
                    col = (0,0,255) if es_neg else (0,255,120)
                    x1,y1,x2,y2 = map(int, box.xyxy[0])
                    cv2.rectangle(frame,(x1,y1),(x2,y2),col,2)
                    cv2.putText(frame, model_epp.names[cls],
                                (x1,y1-5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, col, 1)

        casco_ok   = est_casco.actualizar(casco_crudo)
        chaleco_ok = est_chaleco.actualizar(chaleco_crudo)
        epp_ok     = set()
        if casco_ok:   epp_ok.add("CASCO")
        if chaleco_ok: epp_ok.add("CHALECO")
        epp_falta  = EPP_REQUERIDO - epp_ok

        ahora = time.time()
        if persona_det and epp_falta:
            for epp in sorted(epp_falta):
                if epp not in t_epp_falta:
                    t_epp_falta[epp] = ahora
                elif ahora - t_epp_falta[epp] >= EPP_TIEMPO_GRACIA:
                    if epp not in alarmas_epp:
                        guardar_evento(f"EPP_FALTANTE:{epp}")
                        alarmas_epp.add(epp)
                        threading.Thread(target=alarma_sonido, daemon=True).start()
                        threading.Thread(
                            target=_grabar_clip,
                            args=(0, f"EPP_FALTANTE:{epp}", list(_buf1)),
                            daemon=True).start()
            for epp in list(t_epp_falta):
                if epp not in epp_falta:
                    del t_epp_falta[epp]
        else:
            t_epp_falta = {}
            alarmas_epp = set()

        cam1["ojos"]    = estado_ojos
        cam1["ear"]     = ear_val
        cam1["persona"] = persona_det
        cam1["telefono"]= telefono_det
        cam1["epp_ok"]  = epp_ok
        cam1["epp_falta"]= epp_falta

        # Codificar frame como JPEG
        _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
        cam1["frame_jpg"] = buf.tobytes()

    cam1["running"] = False


def _loop_cam2():
    global tiempo_inicio_ojos2, alarma_fatiga2, alarma_telefono2
    alarmas_epp2 = set()
    t_epp_falta2 = {}

    from collections import deque
    _buf2 = deque(maxlen=30)

    while cam2["running"]:
        ret, frame = cam2["cap"].read()
        if not ret:
            break

        h, w, _ = frame.shape
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        _buf2.append(frame.copy())

        # Fatiga cam2
        estado_ojos  = "SIN DETECCION"
        ear_val      = 0.0
        if face_mesh is not None:
            results_face = face_mesh.process(rgb)
            if results_face.multi_face_landmarks:
                for lm in results_face.multi_face_landmarks:
                    le = calcular_EAR(LEFT_EYE,  lm.landmark, w, h)
                    re = calcular_EAR(RIGHT_EYE, lm.landmark, w, h)
                    ear_val = (le + re) / 2.0
                    op2 = operarios[1]
                    ced2 = op2["cedula"] if op2 else None
                    if ear_val < EAR_UMBRAL:
                        estado_ojos = "OJOS CERRADOS"
                        if tiempo_inicio_ojos2 is None:
                            tiempo_inicio_ojos2 = time.time()
                        if time.time() - tiempo_inicio_ojos2 >= TIEMPO_UMBRAL:
                            estado_ojos = "!! FATIGA DETECTADA !!"
                            if not alarma_fatiga2:
                                threading.Thread(target=alarma_sonido, daemon=True).start()
                                guardar_evento("FATIGA", cedula=ced2)
                                alarma_fatiga2 = True
                                threading.Thread(
                                    target=_grabar_clip,
                                    args=(1, "FATIGA", list(_buf2)),
                                    daemon=True).start()
                    else:
                        estado_ojos = "ACTIVO"
                        tiempo_inicio_ojos2 = None
                        alarma_fatiga2 = False

        # YOLO general cam2
        persona_cruda  = False
        telefono_crudo = False
        if YOLO_OK and model_general:
            op2 = operarios[1]
            ced2 = op2["cedula"] if op2 else None
            for r in model_general(frame, verbose=False):
                for box in r.boxes:
                    cls = int(box.cls[0])
                    nm  = model_general.names[cls].lower()
                    if cls == 0: persona_cruda = True
                    if "cell phone" in nm or "phone" in nm: telefono_crudo = True
                    x1,y1,x2,y2 = map(int, box.xyxy[0])
                    cv2.rectangle(frame,(x1,y1),(x2,y2),(255,220,0),2)
                    cv2.putText(frame, model_general.names[cls],
                                (x1,y1-5), cv2.FONT_HERSHEY_SIMPLEX, 0.5,(255,220,0),1)

            persona_det  = est2_persona.actualizar(persona_cruda)
            telefono_det = est2_telefono.actualizar(telefono_crudo)
            if telefono_det and not alarma_telefono2:
                threading.Thread(target=alarma_sonido, daemon=True).start()
                guardar_evento("TELEFONO", cedula=ced2)
                alarma_telefono2 = True
                threading.Thread(
                    target=_grabar_clip,
                    args=(1, "TELEFONO", list(_buf2)),
                    daemon=True).start()
            elif not telefono_det:
                alarma_telefono2 = False
        else:
            persona_det  = est2_persona.actualizar(persona_cruda)
            telefono_det = est2_telefono.actualizar(telefono_crudo)

        # YOLO EPP cam2
        casco_crudo = chaleco_crudo = False
        if YOLO_OK and model_epp:
            op2 = operarios[1]
            ced2 = op2["cedula"] if op2 else None
            for r in model_epp(frame, verbose=False):
                for box in r.boxes:
                    cls = int(box.cls[0])
                    nm  = model_epp.names[cls].lower()
                    if nm == "hardhat":     casco_crudo   = True
                    if nm == "safety vest": chaleco_crudo = True
                    es_neg = any(k in nm for k in EPP_NEGATIVO)
                    col = (0,0,255) if es_neg else (0,255,120)
                    x1,y1,x2,y2 = map(int, box.xyxy[0])
                    cv2.rectangle(frame,(x1,y1),(x2,y2),col,2)
                    cv2.putText(frame, model_epp.names[cls],
                                (x1,y1-5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, col, 1)

        casco_ok   = est2_casco.actualizar(casco_crudo)
        chaleco_ok = est2_chaleco.actualizar(chaleco_crudo)
        epp_ok     = set()
        if casco_ok:   epp_ok.add("CASCO")
        if chaleco_ok: epp_ok.add("CHALECO")
        epp_falta  = EPP_REQUERIDO - epp_ok

        ahora = time.time()
        op2 = operarios[1]
        ced2 = op2["cedula"] if op2 else None
        if persona_det and epp_falta:
            for epp in sorted(epp_falta):
                if epp not in t_epp_falta2:
                    t_epp_falta2[epp] = ahora
                elif ahora - t_epp_falta2[epp] >= EPP_TIEMPO_GRACIA:
                    if epp not in alarmas_epp2:
                        guardar_evento(f"EPP_FALTANTE:{epp}", cedula=ced2)
                        alarmas_epp2.add(epp)
                        threading.Thread(target=alarma_sonido, daemon=True).start()
                        threading.Thread(
                            target=_grabar_clip,
                            args=(1, f"EPP_FALTANTE:{epp}", list(_buf2)),
                            daemon=True).start()
            for epp in list(t_epp_falta2):
                if epp not in epp_falta:
                    del t_epp_falta2[epp]
        else:
            t_epp_falta2 = {}
            alarmas_epp2 = set()

        cam2["ojos"]     = estado_ojos
        cam2["ear"]      = ear_val
        cam2["persona"]  = persona_det
        cam2["telefono"] = telefono_det
        cam2["epp_ok"]   = epp_ok
        cam2["epp_falta"]= epp_falta

        _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
        cam2["frame_jpg"] = buf.tobytes()

    cam2["running"] = False


# ── Generadores MJPEG ─────────────────────────────────────────
def gen_cam1():
    while True:
        jpg = cam1.get("frame_jpg")
        if jpg:
            yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + jpg + b"\r\n")
        time.sleep(0.033)

def gen_cam2():
    while True:
        jpg = cam2.get("frame_jpg")
        if jpg:
            yield (b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + jpg + b"\r\n")
        time.sleep(0.033)


# ── Exportar Excel ────────────────────────────────────────────
def _exportar_excel_registros():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, "openpyxl no instalado"
    registros = cargar_registros()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros Operarios"
    headers = ["Cédula","Nombre","Cargo","Entrada","Salida","Duración (min)"]
    hf = PatternFill("solid", fgColor="FF6B00")
    hft = Font(bold=True, color="FFFFFF", name="Calibri")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hft
        cell.alignment = Alignment(horizontal="center")
    for ri, r in enumerate(registros, 2):
        for ci, key in enumerate(["Cedula","Nombre","Cargo","Entrada","Salida","Duracion_min"], 1):
            ws.cell(row=ri, column=ci, value=r.get(key,""))
    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = mx + 4
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name, None

def _exportar_excel_eventos():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, "openpyxl no instalado"
    eventos = cargar_eventos()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eventos"
    headers = ["Fecha/Hora","Cédula","Evento"]
    hf = PatternFill("solid", fgColor="FF6B00")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFF")
    for ri, ev in enumerate(eventos, 2):
        ws.cell(row=ri, column=1, value=ev["ts"])
        ws.cell(row=ri, column=2, value=ev["cedula"])
        ws.cell(row=ri, column=3, value=ev["tipo"])
    for col in ws.columns:
        mx = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = mx + 4
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name, None


# ── Gráficas base64 ───────────────────────────────────────────
# ── Flask app ─────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "securework_secret_key_2024"
inicializar_csv()


def login_required(rol=None):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "usuario" not in session:
                return jsonify({"error": "No autenticado"}), 401
            if rol and session.get("rol") != rol:
                return jsonify({"error": "Sin permisos"}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


# ── HTML embebido (sin carpeta templates/) ────────────────────

# ── HTML embebido (sin carpeta templates/) ────────────────────
INDEX_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SecureWork — Sistema de Monitoreo</title>
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;600;700&family=Syne:wght@400;600;700;800&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/three.js/r128/three.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/gsap/3.12.5/gsap.min.js"></script>
<style>
:root {
  --bg:#0a0a0f;--panel:#111118;--card:#16161f;--border:#252535;
  --accent:#ff6b00;--accent2:#ff8c00;--text:#eaeaf2;--sub:#7a7a9a;
  --green:#00e676;--red:#ff3d5a;--yellow:#ffca28;--blue:#448aff;
  --radius:10px;--gap:14px;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html,body{height:100%;font-family:'JetBrains Mono',monospace;background:var(--bg);color:var(--text);font-size:13px;overflow:hidden}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:var(--panel)}
::-webkit-scrollbar-thumb{background:var(--border);border-radius:4px}

/* ── FONDO ANIMADO LightPillar ── */
#light-pillar-bg{position:fixed;top:0;left:0;width:100%;height:100%;z-index:0;pointer-events:none;mix-blend-mode:screen;opacity:0.85}

/* ── LOGIN ── */
#login-screen{position:fixed;inset:0;display:flex;align-items:center;justify-content:center;background:transparent;z-index:1000}
.login-card{width:400px;background:rgba(22,22,31,0.92);border:1px solid var(--border);border-radius:16px;overflow:hidden;box-shadow:0 30px 80px rgba(0,0,0,.6),0 0 0 1px rgba(255,107,0,.08);animation:slideUp .4s cubic-bezier(.16,1,.3,1);backdrop-filter:blur(18px)}
@keyframes slideUp{from{transform:translateY(30px);opacity:0}to{transform:translateY(0);opacity:1}}
.login-top-bar{height:3px;background:linear-gradient(90deg,var(--accent),var(--accent2))}
.login-body{padding:40px 36px}
.login-logo{text-align:center;margin-bottom:32px}
.login-logo .hex{font-size:38px;color:var(--accent);line-height:1;display:inline-block;animation:hexSpin 8s linear infinite;transform-origin:center}
.login-logo .brand{font-family:'Syne',sans-serif;font-size:22px;font-weight:800;color:var(--text);letter-spacing:-.5px;margin-top:6px}
.login-logo .sub{color:var(--sub);font-size:11px;margin-top:4px}
.login-label{font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px;display:block}
.login-input{width:100%;background:var(--border);border:1px solid transparent;color:var(--text);font-family:inherit;font-size:13px;padding:11px 14px;border-radius:8px;outline:none;transition:border-color .2s}
.login-input:focus{border-color:var(--accent)}
.login-field{margin-bottom:18px}
.login-error{font-size:11px;color:var(--red);min-height:18px;margin-bottom:10px}
.btn-login{width:100%;padding:13px;background:var(--accent);color:#fff;font-family:inherit;font-size:13px;font-weight:700;letter-spacing:.06em;border:none;border-radius:8px;cursor:pointer;transition:background .15s,transform .1s}
.btn-login:hover{background:var(--accent2)}
.btn-login:active{transform:scale(.98)}

/* ── APP SHELL ── */
#app{height:100vh;flex-direction:column;position:relative;z-index:1;display:none}
#app.visible{display:flex}
.header{height:54px;background:rgba(17,17,24,0.88);border-bottom:1px solid var(--border);display:flex;align-items:center;padding:0 20px;gap:12px;flex-shrink:0;position:relative;backdrop-filter:blur(12px)}
.header::after{content:'';position:absolute;bottom:0;left:0;right:0;height:1px;background:linear-gradient(90deg,var(--accent) 0%,transparent 40%)}
.header-brand{display:flex;align-items:center;gap:10px}
.header-hex{font-size:20px;color:var(--accent);display:inline-block;animation:hexSpin 8s linear infinite;transform-origin:center}
@keyframes hexSpin{0%{transform:rotateY(0deg)}50%{transform:rotateY(180deg)}100%{transform:rotateY(360deg)}}
.header-title{font-family:'Syne',sans-serif;font-weight:800;font-size:16px;letter-spacing:-.3px}
.header-title span{color:var(--accent)}
.header-right{margin-left:auto;display:flex;align-items:center;gap:16px}
.header-clock{color:var(--sub);font-size:11px}
.cam-badge{display:flex;align-items:center;gap:6px;font-size:11px;padding:4px 10px;border-radius:20px;background:var(--card);border:1px solid var(--border)}
.cam-dot{width:7px;height:7px;border-radius:50%;background:var(--red)}
.cam-dot.active{background:var(--green);box-shadow:0 0 6px var(--green),0 0 12px rgba(0,230,118,.3)}
.btn-logout{background:none;border:1px solid var(--border);color:var(--sub);font-family:inherit;font-size:11px;padding:5px 12px;border-radius:6px;cursor:pointer;transition:all .15s}
.btn-logout:hover{border-color:var(--red);color:var(--red)}
.main-layout{display:flex;flex:1;overflow:hidden}
/* ── STAGGERED MENU ── */
.staggered-menu-wrapper{position:relative;--sm-accent:var(--accent)}
.staggered-menu-wrapper[data-position="left"] .sm-prelayers,
.staggered-menu-wrapper[data-position="left"] .staggered-menu-panel{left:0;right:auto}
.sm-prelayers{position:fixed;top:0;left:0;width:260px;height:100vh;z-index:200;pointer-events:none}
.sm-prelayer{position:absolute;inset:0;will-change:transform}
.staggered-menu-panel{position:fixed;top:0;left:0;width:260px;height:100vh;background:var(--panel);z-index:210;display:flex;flex-direction:column;will-change:transform;overflow-y:auto;border-right:1px solid var(--border)}
.sm-panel-inner{padding:80px 32px 32px;display:flex;flex-direction:column;height:100%}
.sm-panel-list{list-style:none;margin:0;padding:0;flex:1}
.sm-panel-itemWrap{overflow:hidden;margin-bottom:4px}
.sm-panel-item{display:flex;align-items:center;gap:10px;padding:10px 0;text-decoration:none;color:var(--text);cursor:pointer;background:none;border:none;font-family:inherit;width:100%;position:relative}
.sm-panel-item::before{content:attr(data-emoji);font-size:18px;opacity:var(--sm-num-opacity,0);min-width:28px;transition:opacity .2s;line-height:1}
.sm-panel-list[data-numbering] .sm-panel-item{--sm-num-opacity:0}
.sm-panel-itemLabel{font-family:'JetBrains Mono',monospace;font-size:20px;font-weight:700;display:block;will-change:transform;transition:color .2s;letter-spacing:-.5px}
.sm-panel-item:hover .sm-panel-itemLabel{color:var(--accent)}
.sm-socials{margin-top:auto;padding-top:24px;border-top:1px solid var(--border)}
.sm-socials-title{font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.1em;margin-bottom:12px;font-weight:700}
.sm-socials-list{list-style:none;padding:0;margin:0;display:flex;gap:16px;flex-wrap:wrap}
.sm-socials-link{color:var(--sub);text-decoration:none;font-size:11px;font-weight:600;transition:color .2s}
.sm-socials-link:hover{color:var(--accent)}
/* Toggle button */
.sm-toggle{background:none;border:none;cursor:pointer;display:flex;align-items:center;gap:8px;padding:6px 10px;border-radius:8px;transition:background .15s;color:var(--accent);font-family:inherit}
.sm-toggle:hover{background:rgba(255,107,0,.08)}
.sm-toggle-textWrap{overflow:hidden;height:1.1em;display:inline-block;vertical-align:middle}
.sm-toggle-textInner{display:flex;flex-direction:column;will-change:transform}
.sm-toggle-line{font-size:11px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;line-height:1.1em;white-space:nowrap}
.sm-icon{width:18px;height:18px;position:relative;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0}
.sm-icon-line{position:absolute;width:100%;height:2px;background:currentColor;border-radius:2px;will-change:transform}
.sm-icon-line-v{width:2px;height:100%}
.sidebar-actions{padding:12px 14px;display:flex;flex-direction:column;gap:6px}
.btn-entrada{background:rgba(0,230,118,.12);color:var(--green);border:1px solid rgba(0,230,118,.25);font-family:inherit;font-size:11px;padding:8px;border-radius:8px;cursor:pointer;font-weight:600;transition:all .15s;width:100%}
.btn-entrada:hover{background:rgba(0,230,118,.2)}
.btn-salida{background:rgba(255,61,90,.12);color:var(--red);border:1px solid rgba(255,61,90,.25);font-family:inherit;font-size:11px;padding:8px;border-radius:8px;cursor:pointer;font-weight:600;transition:all .15s;width:100%}
.btn-salida:hover{background:rgba(255,61,90,.2)}
/* Sidebar-actions e operario dentro del panel */
.sm-panel-actions{display:flex;flex-direction:column;gap:6px;margin-top:16px}
.sm-panel-operario{padding:14px 0 0;border-top:1px solid var(--border);margin-top:12px}
.sidebar{display:none}/* ocultar sidebar vieja - ahora usamos StaggeredMenu */
.content{flex:1;overflow-y:auto;padding:20px}
.page{display:none}
.page.active{display:block}
.page-title{font-family:'Syne',sans-serif;font-size:18px;font-weight:800;color:var(--text);margin-bottom:18px;display:flex;align-items:center;gap:10px;justify-content:space-between}
.page-title-left{display:flex;align-items:center;gap:10px}
.page-title-left span:first-child{color:var(--accent)}
.card{background:rgba(22,22,31,0.85);border:1px solid var(--border);border-radius:var(--radius);backdrop-filter:blur(8px)}
.card-sm{padding:14px 16px}

/* ── CÁMARAS ── */
.cams-grid{display:grid;grid-template-columns:1fr 1fr;gap:var(--gap);margin-bottom:var(--gap)}
.cam-label{font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px;display:flex;align-items:center;gap:8px}
.cam-label .dot{width:6px;height:6px;border-radius:50%;background:var(--red)}
.cam-label .dot.on{background:var(--green);box-shadow:0 0 5px var(--green)}
.cam-feed{width:100%;aspect-ratio:16/9;background:#000;border-radius:var(--radius);border:1px solid var(--border);overflow:hidden;position:relative;display:flex;align-items:center;justify-content:center}
.cam-feed img{width:100%;height:100%;object-fit:cover}
.cam-placeholder{color:var(--sub);font-size:12px;text-align:center;line-height:1.8}
.cam-placeholder .icon{font-size:28px;margin-bottom:8px;display:block}
.status-grid{display:grid;grid-template-columns:repeat(6,1fr);gap:8px;margin-bottom:var(--gap)}
.status-card{background:rgba(22,22,31,0.85);border:1px solid var(--border);border-radius:var(--radius);padding:12px 14px}
.status-card .s-label{font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px}
.status-card .s-val{font-size:13px;font-weight:700;color:var(--text)}
.s-ok{color:var(--green)!important}
.s-warn{color:var(--red)!important}
.s-yellow{color:var(--yellow)!important}
.s-blue{color:var(--blue)!important}
.alerts-bar{background:rgba(22,22,31,0.85);border:1px solid var(--border);border-radius:var(--radius);padding:12px 16px;display:flex;flex-wrap:wrap;gap:10px;align-items:center;min-height:44px}
.alert-chip{display:flex;align-items:center;gap:6px;padding:5px 12px;border-radius:20px;font-size:11px;font-weight:700;animation:pulseIn .3s ease}
@keyframes pulseIn{from{transform:scale(.9);opacity:0}to{transform:scale(1);opacity:1}}
.alert-chip.red{background:rgba(255,61,90,.15);color:var(--red);border:1px solid rgba(255,61,90,.3)}
.alert-chip.yellow{background:rgba(255,202,40,.12);color:var(--yellow);border:1px solid rgba(255,202,40,.25)}
.alert-chip.orange{background:rgba(255,107,0,.15);color:var(--accent);border:1px solid rgba(255,107,0,.3)}
.no-alerts{color:var(--green);font-size:12px;display:flex;align-items:center;gap:8px}
.cam-controls{display:flex;gap:8px;margin-bottom:var(--gap)}
.btn-primary{background:var(--accent);color:#fff;border:none;font-family:inherit;font-size:12px;font-weight:700;padding:9px 20px;border-radius:8px;cursor:pointer;transition:background .15s,transform .1s;display:flex;align-items:center;gap:6px}
.btn-primary:hover{background:var(--accent2)}
.btn-primary:active{transform:scale(.97)}
.btn-secondary{background:var(--card);color:var(--text);border:1px solid var(--border);font-family:inherit;font-size:12px;padding:9px 16px;border-radius:8px;cursor:pointer;transition:all .15s;display:flex;align-items:center;gap:6px}
.btn-secondary:hover{border-color:var(--sub)}
.btn-danger{background:var(--red);color:#fff;border:none;font-family:inherit;font-size:12px;font-weight:700;padding:9px 20px;border-radius:8px;cursor:pointer;transition:background .15s}
.btn-danger:hover{background:#cc2040}

/* ── TABLA ── */
.tbl-wrap{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:12px}
thead th{background:var(--border);color:var(--accent);font-weight:700;padding:10px 14px;text-align:left;white-space:nowrap;text-transform:uppercase;font-size:10px;letter-spacing:.06em}
tbody tr{border-bottom:1px solid var(--border);transition:background .1s}
tbody tr:hover{background:rgba(255,255,255,.025)}
tbody td{padding:10px 14px;color:var(--text)}
.tag{display:inline-flex;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:700}
.tag-red{background:rgba(255,61,90,.15);color:var(--red)}
.tag-yellow{background:rgba(255,202,40,.12);color:var(--yellow)}
.tag-green{background:rgba(0,230,118,.12);color:var(--green)}
.tag-blue{background:rgba(68,138,255,.12);color:var(--blue)}
.tag-orange{background:rgba(255,107,0,.15);color:var(--accent)}
.tag-gray{background:rgba(122,122,154,.12);color:var(--sub)}

/* ── KPI CARDS ── */
.kpi-row{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px}
.kpi-card{background:rgba(22,22,31,0.85);border:1px solid var(--border);border-radius:var(--radius);padding:16px;text-align:center;position:relative;overflow:hidden;transition:transform .15s,border-color .15s;backdrop-filter:blur(8px)}
.kpi-card:hover{transform:translateY(-2px);border-color:var(--accent)}
.kpi-card-bar{position:absolute;top:0;left:0;right:0;height:3px}
.kpi-val{font-family:'Syne',sans-serif;font-size:28px;font-weight:800;line-height:1}
.kpi-lbl{font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-top:6px}

/* ── GRÁFICAS CHARTJS ── */
.charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:16px}
.chart-full{grid-column:1/-1}
.chart-card{background:rgba(22,22,31,0.88);border:1px solid var(--border);border-radius:var(--radius);padding:16px;overflow:hidden;transition:border-color .15s;backdrop-filter:blur(8px)}
.chart-card:hover{border-color:rgba(255,107,0,.35)}
.chart-card-title{font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.08em;margin-bottom:12px;font-weight:700}
.chart-container{position:relative;width:100%}
.chart-container canvas{width:100%!important}
.chart-empty{color:var(--sub);font-size:12px;text-align:center;padding:50px 20px;display:flex;flex-direction:column;gap:8px;align-items:center}
.chart-empty::before{content:'◈';font-size:24px;opacity:.25}

/* ── PERIODO TABS ── */
.period-tabs{display:flex;gap:6px;margin-bottom:14px;flex-wrap:wrap;align-items:center}
.tab-btn{background:var(--border);color:var(--sub);border:none;font-family:inherit;font-size:11px;padding:6px 14px;border-radius:6px;cursor:pointer;transition:all .15s;font-weight:600}
.tab-btn.active{background:var(--accent);color:#fff}
.tab-btn:hover:not(.active){background:var(--card);color:var(--text)}

/* ── MODAL ── */
.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.7);backdrop-filter:blur(4px);display:none;align-items:center;justify-content:center;z-index:500}
.modal-overlay.open{display:flex}
.modal{background:var(--card);border:1px solid var(--border);border-radius:14px;width:400px;overflow:hidden;animation:slideUp .3s cubic-bezier(.16,1,.3,1);box-shadow:0 30px 70px rgba(0,0,0,.5)}
.modal-top{height:3px;background:linear-gradient(90deg,var(--accent),var(--accent2))}
.modal-body{padding:28px 28px 24px}
.modal-title{font-family:'Syne',sans-serif;font-size:16px;font-weight:800;color:var(--accent);margin-bottom:20px}
.field-label{font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px;display:block}
.field-input{width:100%;background:var(--border);border:1px solid transparent;color:var(--text);font-family:inherit;font-size:12px;padding:10px 12px;border-radius:7px;outline:none;margin-bottom:14px;transition:border-color .15s}
.field-input:focus{border-color:var(--accent)}
.modal-actions{display:flex;gap:8px;justify-content:flex-end;margin-top:8px}
.btn-confirm{background:var(--accent);color:#fff;border:none;font-family:inherit;font-size:12px;font-weight:700;padding:10px 22px;border-radius:8px;cursor:pointer}
.btn-cancel{background:var(--border);color:var(--sub);border:none;font-family:inherit;font-size:12px;padding:10px 16px;border-radius:8px;cursor:pointer}
.modal-err{font-size:11px;color:var(--red);min-height:16px;margin-bottom:8px}

/* ── TOAST ── */
#toast{position:fixed;bottom:24px;right:24px;padding:12px 20px;background:var(--card);border:1px solid var(--border);border-radius:10px;font-size:12px;font-weight:600;color:var(--text);box-shadow:0 10px 40px rgba(0,0,0,.4);z-index:999;transform:translateY(20px);opacity:0;transition:all .25s;pointer-events:none}
#toast.show{transform:translateY(0);opacity:1}
#toast.ok{border-left:3px solid var(--green)}
#toast.err{border-left:3px solid var(--red)}

/* ── CONFIGURACIÓN ── */
.config-section{margin-bottom:24px}
.config-section-title{font-size:11px;color:var(--accent);text-transform:uppercase;letter-spacing:.08em;font-weight:700;padding-bottom:10px;border-bottom:1px solid var(--border);margin-bottom:14px}
.slider-row{background:rgba(22,22,31,0.85);border:1px solid var(--border);border-radius:8px;padding:14px 16px;margin-bottom:8px;display:flex;align-items:center;gap:12px}
.slider-row label{flex:1;font-size:12px;color:var(--text)}
.slider-row .hint{font-size:10px;color:var(--sub);margin-top:3px}
.slider-row input[type=range]{width:180px;accent-color:var(--accent)}
.slider-val{width:48px;text-align:right;font-size:13px;font-weight:700;color:var(--accent)}
.checkbox-row{display:flex;gap:20px;padding:12px 0}
.checkbox-row label{display:flex;align-items:center;gap:8px;cursor:pointer;font-size:12px}
.checkbox-row input{accent-color:var(--accent);width:16px;height:16px}

/* ── PORTAL EMPLEADO ── */
.emp-portal{display:flex;flex-direction:column;align-items:center;min-height:calc(100vh - 54px);gap:20px;padding:24px;overflow-y:auto}
.emp-card{background:rgba(22,22,31,0.85);border:1px solid var(--border);border-radius:14px;padding:28px;width:100%;max-width:460px;backdrop-filter:blur(8px)}
.emp-card-title{font-size:11px;color:var(--sub);text-transform:uppercase;letter-spacing:.08em;margin-bottom:12px}
.emp-status{font-size:14px;font-weight:700;color:var(--sub)}
.emp-status.active{color:var(--green)}
.emp-btns{display:flex;flex-direction:column;gap:10px;width:100%;max-width:460px}
.emp-btn{padding:14px;border-radius:10px;font-family:inherit;font-size:13px;font-weight:700;cursor:pointer;border:none;transition:all .15s;display:flex;align-items:center;justify-content:center;gap:10px}
.emp-btn.green{background:rgba(0,230,118,.12);color:var(--green);border:1px solid rgba(0,230,118,.25)}
.emp-btn.green:hover{background:rgba(0,230,118,.22)}
.emp-btn.red{background:rgba(255,61,90,.12);color:var(--red);border:1px solid rgba(255,61,90,.25)}
.emp-btn.red:hover{background:rgba(255,61,90,.22)}
.stats-emp-search{display:flex;gap:8px;margin-bottom:14px}
.stats-emp-search input{flex:1;background:var(--card);border:1px solid var(--border);color:var(--text);font-family:inherit;font-size:12px;padding:9px 12px;border-radius:8px;outline:none}
.stats-emp-search input:focus{border-color:var(--accent)}
.sep{border:none;border-top:1px solid var(--border);margin:16px 0}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.3}}
@media(max-width:900px){
  .cams-grid{grid-template-columns:1fr}
  .kpi-row{grid-template-columns:repeat(2,1fr)}
  .charts-grid{grid-template-columns:1fr}
  .chart-full{grid-column:1}
}
</style>
</head>
<body>

<!-- FONDO ANIMADO LightPillar -->
<div id="light-pillar-bg"></div>

<!-- LOGIN -->
<div id="login-screen">
  <div class="login-card">
    <div class="login-top-bar"></div>
    <div class="login-body">
      <div class="login-logo">
        <div class="hex">⬡</div>
        <div class="brand">SecureWork</div>
        <div class="sub">Sistema de Monitoreo de Seguridad</div>
      </div>
      <div class="login-field">
        <span class="login-label">Usuario</span>
        <input class="login-input" id="l-user" type="text" placeholder="admin / empleado" autocomplete="off">
      </div>
      <div class="login-field">
        <span class="login-label">Contraseña</span>
        <input class="login-input" id="l-pass" type="password" placeholder="••••••••">
      </div>
      <div id="modelos-status" style="display:flex;align-items:center;gap:8px;background:rgba(255,107,0,.08);border:1px solid rgba(255,107,0,.2);border-radius:8px;padding:10px 14px;margin-bottom:12px;font-size:11px;color:var(--accent)">
        <span id="modelos-dot" style="width:8px;height:8px;border-radius:50%;background:var(--accent);animation:blink 1s infinite;flex-shrink:0"></span>
        <span id="modelos-txt">Cargando modelos de IA en segundo plano (puedes ingresar ya)...</span>
      </div>
      <div class="login-error" id="l-err"></div>
      <button class="btn-login" id="btn-login">INGRESAR →</button>
    </div>
  </div>
</div>

<!-- APP -->
<div id="app">
  <div id="banner-modelos" style="display:none;align-items:center;gap:10px;background:rgba(255,107,0,.10);border-bottom:1px solid rgba(255,107,0,.25);padding:8px 20px;font-size:11px;color:var(--accent);flex-shrink:0;">
    <span style="width:7px;height:7px;border-radius:50%;background:var(--accent);animation:blink 1s infinite;flex-shrink:0;display:inline-block"></span>
    <span>Modelos de IA cargando en segundo plano — la detección estará disponible en unos momentos.</span>
  </div>

  <header class="header">
    <div class="header-brand">
      <span class="header-hex">⬡</span>
      <span class="header-title">Secure<span>Work</span></span>
    </div>
    <div class="header-right">
      <span class="header-clock" id="hdr-clock"></span>
      <div class="cam-badge" id="cam-badge">
        <span class="cam-dot" id="cam-dot"></span>
        <span id="cam-badge-txt">SIN CÁMARA</span>
      </div>
      <!-- Botón toggle menú -->
      <button id="sm-toggle-btn" class="sm-toggle" aria-label="Abrir menú" aria-expanded="false" aria-controls="staggered-menu-panel">
        <span id="sm-text-wrap" class="sm-toggle-textWrap" aria-hidden="true">
          <span id="sm-text-inner" class="sm-toggle-textInner">
            <span class="sm-toggle-line">Menú</span>
          </span>
        </span>
        <span id="sm-icon" class="sm-icon" aria-hidden="true">
          <span id="sm-plus-h" class="sm-icon-line"></span>
          <span id="sm-plus-v" class="sm-icon-line sm-icon-line-v"></span>
        </span>
      </button>
      <button class="btn-logout" id="btn-logout">↩ SALIR</button>
    </div>
  </header>

  <!-- STAGGERED MENU -->
  <div id="staggered-menu-wrapper" class="staggered-menu-wrapper" data-position="left">
    <div id="sm-prelayers" class="sm-prelayers" aria-hidden="true">
      <div class="sm-prelayer" style="background:#ff8c00"></div>
      <div class="sm-prelayer" style="background:#ff6b00"></div>
    </div>
    <aside id="staggered-menu-panel" class="staggered-menu-panel" aria-hidden="true">
      <div class="sm-panel-inner">
        <!-- Logo -->
        <div style="margin-bottom:28px;padding-bottom:20px;border-bottom:1px solid var(--border)">
          <div style="display:flex;align-items:center;gap:10px">
            <span style="font-size:26px;color:var(--accent);display:inline-block;animation:hexSpin 8s linear infinite;transform-origin:center">⬡</span>
            <span style="font-family:'Syne',sans-serif;font-weight:800;font-size:18px;letter-spacing:-.3px">Secure<span style="color:var(--accent)">Work</span></span>
          </div>
          <div style="font-size:10px;color:var(--sub);margin-top:4px;letter-spacing:.06em">Sistema de Monitoreo de Seguridad</div>
        </div>
        <ul class="sm-panel-list" role="list" data-numbering="true" id="sm-nav-list">
          <li class="sm-panel-itemWrap"><button class="sm-panel-item nav-btn" data-page="camara"><span class="sm-panel-itemLabel">Cámara</span></button></li>
          <li class="sm-panel-itemWrap"><button class="sm-panel-item nav-btn" data-page="registros"><span class="sm-panel-itemLabel">Registros</span></button></li>
          <li class="sm-panel-itemWrap"><button class="sm-panel-item nav-btn" data-page="eventos"><span class="sm-panel-itemLabel">Eventos</span></button></li>
          <li class="sm-panel-itemWrap"><button class="sm-panel-item nav-btn" data-page="estadisticas"><span class="sm-panel-itemLabel">Estadísticas</span></button></li>
          <li class="sm-panel-itemWrap"><button class="sm-panel-item nav-btn" data-page="evidencias"><span class="sm-panel-itemLabel">Evidencias</span></button></li>
          <li class="sm-panel-itemWrap"><button class="sm-panel-item nav-btn" data-page="configuracion"><span class="sm-panel-itemLabel">Configuración</span></button></li>
        </ul>
        <div class="sidebar-actions" style="padding:12px 0;margin-top:8px">
          <button class="btn-entrada" onclick="openModal('entrada');smClose()">+ REGISTRAR ENTRADA</button>
          <button class="btn-salida"  onclick="openModal('salida');smClose()">— REGISTRAR SALIDA</button>
        </div>
        <div class="sm-panel-operario">
          <div style="font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px;font-weight:700">Operario activo</div>
          <div class="sidebar-op-name none" id="sb-op-name">— Sin registro —</div>
          <div class="sidebar-op-sub" id="sb-op-sub"></div>
        </div>
        <div class="sm-socials">
          <h3 class="sm-socials-title">SecureWork</h3>
          <ul class="sm-socials-list" role="list">
            <li class="sm-socials-item"><span class="sm-socials-link">v2.0 · Monitoreo IA</span></li>
          </ul>
        </div>
      </div>
    </aside>
  </div>

  <div class="main-layout">
    <nav class="sidebar" id="sidebar" style="display:none"></nav>

    <!-- PORTAL EMPLEADO (fuera de main-content para evitar conflictos de display) -->
    <div id="page-empleado" class="emp-portal" style="display:none;flex:1">
      <!-- Tarjeta Operario 1 (Cámara 1) -->
      <div class="emp-card" id="emp-card-op1">
        <div class="emp-card-title">👤 Operario en Cámara 1</div>
        <div class="emp-status none" id="emp-op1-status">— Ninguno registrado —</div>
        <div id="emp-cam1-feed" style="display:none;margin-top:8px">
          <div style="font-size:10px;color:var(--sub);margin-bottom:4px">● CÁMARA 1 ACTIVA</div>
          <img id="emp-feed1" src="" style="width:100%;border-radius:8px;border:1px solid var(--border)">
        </div>
        <div id="emp-cam1-off" style="color:var(--red);font-size:12px;margin-top:8px">● Cámara 1 inactiva</div>
      </div>
      <!-- Tarjeta Operario 2 (Cámara 2) -->
      <div class="emp-card" id="emp-card-op2" style="display:none">
        <div class="emp-card-title">👤 Operario en Cámara 2</div>
        <div class="emp-status none" id="emp-op2-status">— Ninguno registrado —</div>
        <div id="emp-cam2-feed" style="display:none;margin-top:8px">
          <div style="font-size:10px;color:var(--sub);margin-bottom:4px">● CÁMARA 2 ACTIVA</div>
          <img id="emp-feed2" src="" style="width:100%;border-radius:8px;border:1px solid var(--border)">
        </div>
        <div id="emp-cam2-off" style="color:var(--red);font-size:12px;margin-top:8px">● Cámara 2 inactiva</div>
      </div>
      <div class="emp-btns">
        <button class="emp-btn green" onclick="openModal('entrada')">＋ REGISTRAR ENTRADA → INICIAR TURNO</button>
        <button class="emp-btn red"   onclick="openModal('salida')">─ REGISTRAR SALIDA → FINALIZAR TURNO</button>
      </div>
    </div>

    <main class="content" id="main-content">

      <!-- CÁMARAS -->
      <div id="page-camara" class="page active">
        <div class="page-title">
          <div class="page-title-left"><span>📷</span><span>Cámaras en Vivo</span></div>
          <div style="display:flex;gap:8px">
            <button class="btn-primary" id="btn-toggle-cam1" onclick="toggleCam(1)">▶ ENCENDER CÁMARA 1</button>
            <button class="btn-primary" id="btn-toggle-cam2" onclick="toggleCam(2)" style="background:var(--blue)">▶ ENCENDER CÁMARA 2</button>
          </div>
        </div>
        <div class="cams-grid">
          <div class="cam-wrapper">
            <div class="cam-label"><span class="dot" id="dot1"></span>OPERADOR 1 — CÁMARA USB 1 <span id="admin-op1-label" style="margin-left:8px;color:var(--green);font-size:10px"></span></div>
            <div class="cam-feed" id="feed1-wrap">
              <div class="cam-placeholder" id="placeholder1"><span class="icon">📷</span>Cámara inactiva<br>Presiona ENCENDER CÁMARA 1</div>
              <img id="feed1" src="" alt="" style="display:none">
            </div>
          </div>
          <div class="cam-wrapper">
            <div class="cam-label"><span class="dot" id="dot2"></span>OPERADOR 2 — CÁMARA USB 2 <span id="admin-op2-label" style="margin-left:8px;color:var(--green);font-size:10px"></span></div>
            <div class="cam-feed" id="feed2-wrap">
              <div class="cam-placeholder" id="placeholder2"><span class="icon">📷</span>Cámara inactiva<br>Presiona ENCENDER CÁMARA 2</div>
              <img id="feed2" src="" alt="" style="display:none">
            </div>
          </div>
        </div>
        <div class="status-grid">
          <div class="status-card"><div class="s-label">Estado Ojos — CAM1</div><div class="s-val" id="st-ojos">SIN DETECCIÓN</div></div>
          <div class="status-card"><div class="s-label">EAR — CAM1</div><div class="s-val" id="st-ear">0.00</div></div>
          <div class="status-card"><div class="s-label">Persona — CAM1</div><div class="s-val" id="st-persona">NO DETECTADA</div></div>
          <div class="status-card"><div class="s-label">Casco — CAM1</div><div class="s-val s-warn" id="st-casco">[✗] FALTANTE</div></div>
          <div class="status-card"><div class="s-label">Chaleco — CAM1</div><div class="s-val s-warn" id="st-chaleco">[✗] FALTANTE</div></div>
          <div class="status-card"><div class="s-label">Teléfono — CAM1</div><div class="s-val s-ok" id="st-tel">SIN USO</div></div>
          <div class="status-card"><div class="s-label">Estado Ojos — CAM2</div><div class="s-val" id="st-ojos2">SIN DETECCIÓN</div></div>
          <div class="status-card"><div class="s-label">EAR — CAM2</div><div class="s-val" id="st-ear2">0.00</div></div>
          <div class="status-card"><div class="s-label">Persona — CAM2</div><div class="s-val" id="st-persona2">NO DETECTADA</div></div>
          <div class="status-card"><div class="s-label">Casco — CAM2</div><div class="s-val s-warn" id="st-casco2">[✗] FALTANTE</div></div>
          <div class="status-card"><div class="s-label">Chaleco — CAM2</div><div class="s-val s-warn" id="st-chaleco2">[✗] FALTANTE</div></div>
          <div class="status-card"><div class="s-label">Teléfono — CAM2</div><div class="s-val s-ok" id="st-tel2">SIN USO</div></div>
        </div>
        <div class="alerts-bar" id="alerts-bar"><div class="no-alerts">✓ Sin alertas activas</div></div>
      </div>

      <!-- REGISTROS -->
      <div id="page-registros" class="page">
        <div class="page-title">
          <div class="page-title-left"><span>☰</span><span>Registros de Operarios</span></div>
          <div style="display:flex;gap:8px">
            <button class="btn-secondary" onclick="loadRegistros()">↻ Actualizar</button>
            <button class="btn-primary" onclick="exportarExcel('registros')">⬇ Excel</button>
          </div>
        </div>
        <div class="card">
          <div class="tbl-wrap">
            <table id="tbl-registros">
              <thead><tr><th>Cédula</th><th>Nombre</th><th>Cargo</th><th>Entrada</th><th>Salida</th><th>Duración (min)</th></tr></thead>
              <tbody id="tbody-registros"></tbody>
            </table>
          </div>
        </div>
      </div>

      <!-- EVENTOS -->
      <div id="page-eventos" class="page">
        <div class="page-title">
          <div class="page-title-left"><span>⚡</span><span>Historial de Eventos</span></div>
          <div style="display:flex;gap:8px">
            <button class="btn-secondary" onclick="loadEventos()">↻ Actualizar</button>
            <button class="btn-primary" onclick="exportarExcel('eventos')">⬇ Excel</button>
          </div>
        </div>
        <div class="card">
          <div class="tbl-wrap">
            <table id="tbl-eventos">
              <thead><tr><th>Fecha / Hora</th><th>Cédula</th><th>Evento</th></tr></thead>
              <tbody id="tbody-eventos"></tbody>
            </table>
          </div>
        </div>
      </div>

      <!-- ESTADÍSTICAS -->
      <div id="page-estadisticas" class="page">
        <div class="page-title">
          <div class="page-title-left"><span>◈</span><span>Estadísticas</span></div>
          <button class="btn-primary" onclick="exportarPDF()">⬇ PDF + ISO 45001</button>
        </div>
        <div style="display:flex;gap:8px;margin-bottom:16px">
          <button class="tab-btn active" id="tab-general-btn" onclick="switchStatsTab('general')">◈ General</button>
          <button class="tab-btn" id="tab-emp-btn" onclick="switchStatsTab('empleado')">👤 Por Empleado</button>
        </div>
        <div class="period-tabs" id="period-tabs">
          <span style="color:var(--sub);font-size:11px;margin-right:4px">Período:</span>
          <button class="tab-btn active" data-periodo="Todo">Todo</button>
          <button class="tab-btn" data-periodo="Día">Hoy</button>
          <button class="tab-btn" data-periodo="Semana">Semana</button>
          <button class="tab-btn" data-periodo="Mes">Mes</button>
          <button class="tab-btn" data-periodo="Año">Año</button>
        </div>
        <div id="date-filter-bar" style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;background:rgba(22,22,31,0.85);border:1px solid var(--border);border-radius:8px;padding:10px 14px;margin-bottom:14px;font-size:12px">
          <span style="color:var(--sub)">📅 Fecha exacta:</span>
          <div style="display:flex;align-items:center;gap:6px">
            <label style="color:var(--sub);font-size:11px">Desde</label>
            <input type="date" id="fecha-desde" style="background:var(--panel);border:1px solid var(--border);color:var(--text);border-radius:6px;padding:4px 8px;font-size:12px;cursor:pointer" onchange="onFechaChange()">
          </div>
          <div style="display:flex;align-items:center;gap:6px">
            <label style="color:var(--sub);font-size:11px">Hasta</label>
            <input type="date" id="fecha-hasta" style="background:var(--panel);border:1px solid var(--border);color:var(--text);border-radius:6px;padding:4px 8px;font-size:12px;cursor:pointer" onchange="onFechaChange()">
          </div>
          <button class="btn-primary" style="padding:4px 12px;font-size:11px" onclick="aplicarFiltroFecha()">🔍 Aplicar</button>
          <button class="btn-secondary" style="padding:4px 12px;font-size:11px" onclick="limpiarFiltroFecha()">✕ Limpiar fechas</button>
          <span id="fecha-label" style="color:var(--accent);font-size:11px;font-weight:700"></span>
        </div>
        <div id="stats-general">
          <div class="kpi-row" id="kpi-row"></div>
          <div class="charts-grid" id="charts-row">
            <div class="chart-card chart-full"><div class="chart-empty">Cargando gráfica...</div></div>
            <div class="chart-card"><div class="chart-empty">Cargando gráfica...</div></div>
            <div class="chart-card"><div class="chart-empty">Cargando gráfica...</div></div>
          </div>
        </div>
        <div id="stats-empleado" style="display:none">
          <div class="stats-emp-search">
            <div style="position:relative;flex:1">
              <input type="text" id="emp-search-input" placeholder="Buscar por nombre o cédula..." onkeydown="if(event.key==='Enter')buscarEmpleado()" oninput="filtrarSugerencias(this.value)" autocomplete="off" style="width:100%">
              <div id="emp-suggestions" style="display:none;position:absolute;top:100%;left:0;right:0;background:var(--card);border:1px solid var(--border);border-top:none;border-radius:0 0 8px 8px;z-index:100;max-height:200px;overflow-y:auto"></div>
            </div>
            <button class="btn-primary" onclick="buscarEmpleado()">Buscar</button>
            <button class="btn-secondary" onclick="clearEmpSearch()">Limpiar</button>
          </div>
          <div id="emp-results">
            <div style="color:var(--sub);font-size:12px;text-align:center;padding:40px">Ingresa el nombre o cédula de un operario para ver su desempeño.</div>
          </div>
        </div>
      </div>

      <!-- EVIDENCIAS -->
      <div id="page-evidencias" class="page">
        <div class="page-title">
          <div class="page-title-left"><span>🎥</span><span>Evidencias en Video</span></div>
          <button class="btn-primary" onclick="cargarEvidencias()">↻ Actualizar</button>
        </div>
        <div class="config-section" style="margin-bottom:14px">
          <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
            <input type="text" id="ev-search" class="field-input" style="flex:1;min-width:200px;max-width:340px;margin:0" placeholder="Filtrar por nombre o cédula..." onkeydown="if(event.key==='Enter')cargarEvidencias()">
            <button class="btn-primary" onclick="cargarEvidencias()">🔍 Buscar</button>
            <button class="btn-secondary" onclick="document.getElementById('ev-search').value='';cargarEvidencias()">✕ Limpiar</button>
          </div>
        </div>
        <div id="ev-results" style="overflow-y:auto;flex:1">
          <div style="color:var(--sub);font-size:12px;text-align:center;padding:40px">Presiona Actualizar para cargar las evidencias registradas.</div>
        </div>
      </div>

      <!-- CONFIGURACIÓN -->
      <div id="page-configuracion" class="page">
        <div class="page-title"><div class="page-title-left"><span>✦</span><span>Configuración</span></div></div>
        <div class="config-section">
          <div class="config-section-title">① Índices de Cámaras USB</div>
          <div class="slider-row">
            <div style="flex:1"><div>Cámara Operador 1</div><div class="hint">Índice de la cámara USB asignada al Operador 1</div></div>
            <input type="range" id="cfg-cam1" min="0" max="5" step="1" value="0" oninput="document.getElementById('cfg-cam1-val').textContent=this.value">
            <span class="slider-val" id="cfg-cam1-val">0</span>
          </div>
          <div class="slider-row">
            <div style="flex:1"><div>Cámara Operador 2</div><div class="hint">Índice de la cámara USB asignada al Operador 2</div></div>
            <input type="range" id="cfg-cam2" min="0" max="5" step="1" value="1" oninput="document.getElementById('cfg-cam2-val').textContent=this.value">
            <span class="slider-val" id="cfg-cam2-val">1</span>
          </div>
        </div>
        <div class="config-section">
          <div class="config-section-title">② Confianza del Modelo EPP</div>
          <div class="slider-row">
            <div style="flex:1"><div>Umbral de confianza EPP</div><div class="hint">Recomendado: 0.40 – 0.60</div></div>
            <input type="range" id="cfg-conf" min="0.10" max="0.90" step="0.05" value="0.25" oninput="document.getElementById('cfg-conf-val').textContent=parseFloat(this.value).toFixed(2)">
            <span class="slider-val" id="cfg-conf-val">0.25</span>
          </div>
        </div>
        <div class="config-section">
          <div class="config-section-title">③ Detección de Fatiga Ocular</div>
          <div class="slider-row">
            <div style="flex:1"><div>Umbral EAR (apertura de ojos)</div><div class="hint">Recomendado: 0.20 – 0.25</div></div>
            <input type="range" id="cfg-ear" min="0.10" max="0.40" step="0.01" value="0.25" oninput="document.getElementById('cfg-ear-val').textContent=parseFloat(this.value).toFixed(2)">
            <span class="slider-val" id="cfg-ear-val">0.25</span>
          </div>
          <div class="slider-row">
            <div style="flex:1"><div>Segundos con ojos cerrados para alerta</div><div class="hint">Recomendado: 4 – 6 s</div></div>
            <input type="range" id="cfg-tfat" min="1" max="15" step="0.5" value="5" oninput="document.getElementById('cfg-tfat-val').textContent=parseFloat(this.value).toFixed(1)">
            <span class="slider-val" id="cfg-tfat-val">5.0</span>
          </div>
        </div>
        <div class="config-section">
          <div class="config-section-title">④ Tiempo de Gracia EPP (segundos)</div>
          <div class="slider-row">
            <div style="flex:1"><div>Gracia antes de alarmar</div><div class="hint">Recomendado: 4 s</div></div>
            <input type="range" id="cfg-gracia" min="1" max="15" step="0.5" value="4" oninput="document.getElementById('cfg-gracia-val').textContent=parseFloat(this.value).toFixed(1)">
            <span class="slider-val" id="cfg-gracia-val">4.0</span>
          </div>
        </div>
        <div class="config-section">
          <div class="config-section-title">⑤ EPP Requerido en el Puesto</div>
          <div class="checkbox-row">
            <label><input type="checkbox" id="cfg-casco" checked> Casco obligatorio</label>
            <label><input type="checkbox" id="cfg-chaleco" checked> Chaleco obligatorio</label>
          </div>
        </div>
        <button class="btn-primary" style="margin-top:8px" onclick="guardarConfig()">✓ GUARDAR CONFIGURACIÓN</button>
      </div>

    </main>
  </div>
</div>

<!-- MODALES -->
<div class="modal-overlay" id="modal-entrada">
  <div class="modal"><div class="modal-top"></div>
    <div class="modal-body">
      <div class="modal-title">REGISTRAR ENTRADA</div>
      <span class="field-label">Cédula / ID</span>
      <input class="field-input" id="m-ced" placeholder="Ej: 123456789">
      <span class="field-label">Nombre completo</span>
      <input class="field-input" id="m-nom" placeholder="Ej: Juan Pérez">
      <span class="field-label">Cargo</span>
      <input class="field-input" id="m-car" placeholder="Ej: Operario de línea">
      <div class="modal-err" id="m-ent-err"></div>
      <div class="modal-actions">
        <button class="btn-cancel" onclick="closeModal('entrada')">Cancelar</button>
        <button class="btn-confirm" onclick="confirmarEntrada()">Confirmar</button>
      </div>
    </div>
  </div>
</div>
<div class="modal-overlay" id="modal-salida">
  <div class="modal"><div class="modal-top"></div>
    <div class="modal-body">
      <div class="modal-title">REGISTRAR SALIDA</div>
      <span class="field-label">Cédula / ID del operario que sale</span>
      <input class="field-input" id="m-sal-ced" placeholder="Ej: 123456789">
      <div class="modal-err" id="m-sal-err"></div>
      <div class="modal-actions">
        <button class="btn-cancel" onclick="closeModal('salida')">Cancelar</button>
        <button class="btn-confirm" onclick="confirmarSalida()">Confirmar</button>
      </div>
    </div>
  </div>
</div>
<div id="toast"></div>

<script>
// ────────────────────────────────────────────────
// FONDO ANIMADO — LightPillar (ReactBits port)
// ────────────────────────────────────────────────
(function(){
  const container = document.getElementById('light-pillar-bg');
  if(!container) return;

  // Config — colores naranja/ámbar del branding SecureWork
  const TOP_COLOR    = '#EAB308';  // ámbar
  const BOTTOM_COLOR = '#FF6B00';  // naranja accent
  const INTENSITY    = 1.0;
  const ROTATION_SPD = 0.25;
  const GLOW         = 0.0025;
  const WIDTH        = 3.2;
  const HEIGHT       = 0.38;
  const NOISE        = 0.45;
  const ROTATION_DEG = 0;
  const ITERATIONS   = 64;
  const WAVE_ITER    = 3;
  const STEP_MULT    = 1.0;

  const w = container.clientWidth  || window.innerWidth;
  const h = container.clientHeight || window.innerHeight;

  // WebGL check
  const testC = document.createElement('canvas');
  const testGL = testC.getContext('webgl') || testC.getContext('experimental-webgl');
  if(!testGL){ container.style.background='#0a0a0f'; return; }

  let renderer, material, scene, camera, geometry;
  const mouseRef = new THREE.Vector2(0,0);
  let timeVal = 0;
  let rafId   = null;

  scene    = new THREE.Scene();
  camera   = new THREE.OrthographicCamera(-1,1,1,-1,0,1);

  try {
    renderer = new THREE.WebGLRenderer({antialias:false,alpha:true,powerPreference:'high-performance',stencil:false,depth:false});
  } catch(e){ return; }

  renderer.setSize(w, h);
  renderer.setPixelRatio(Math.min(window.devicePixelRatio, 2));
  container.appendChild(renderer.domElement);

  const parseCol = hex => { const c = new THREE.Color(hex); return new THREE.Vector3(c.r,c.g,c.b); };
  const pillarRotRad = (ROTATION_DEG * Math.PI) / 180;
  const waveSin = Math.sin(0.4);
  const waveCos = Math.cos(0.4);

  const VS = `
    varying vec2 vUv;
    void main(){ vUv = uv; gl_Position = vec4(position,1.0); }
  `;

  const FS = `
    precision highp float;
    uniform float uTime;
    uniform vec2  uResolution;
    uniform vec3  uTopColor;
    uniform vec3  uBottomColor;
    uniform float uIntensity;
    uniform float uGlowAmount;
    uniform float uPillarWidth;
    uniform float uPillarHeight;
    uniform float uNoiseIntensity;
    uniform float uRotCos;
    uniform float uRotSin;
    uniform float uPillarRotCos;
    uniform float uPillarRotSin;
    uniform float uWaveSin;
    uniform float uWaveCos;
    varying vec2 vUv;

    const float STEP_MULT  = ${STEP_MULT.toFixed(1)};
    const int   MAX_ITER   = ${ITERATIONS};
    const int   WAVE_ITER  = ${WAVE_ITER};

    void main(){
      vec2 uv = (vUv * 2.0 - 1.0) * vec2(uResolution.x / uResolution.y, 1.0);
      uv = vec2(uPillarRotCos*uv.x - uPillarRotSin*uv.y,
                uPillarRotSin*uv.x + uPillarRotCos*uv.y);

      vec3 ro = vec3(0.0, 0.0, -10.0);
      vec3 rd = normalize(vec3(uv, 1.0));

      float rotC = uRotCos;
      float rotS = uRotSin;

      vec3 col = vec3(0.0);
      float t = 0.1;

      for(int i = 0; i < MAX_ITER; i++){
        vec3 p = ro + rd * t;
        p.xz = vec2(rotC*p.x - rotS*p.z, rotS*p.x + rotC*p.z);

        vec3 q = p;
        q.y = p.y * uPillarHeight + uTime;

        float freq = 1.0;
        float amp  = 1.0;
        for(int j = 0; j < WAVE_ITER; j++){
          q.xz = vec2(uWaveCos*q.x - uWaveSin*q.z, uWaveSin*q.x + uWaveCos*q.z);
          q += cos(q.zxy * freq - uTime * float(j) * 2.0) * amp;
          freq *= 2.0;
          amp  *= 0.5;
        }

        float d     = length(cos(q.xz)) - 0.2;
        float bound = length(p.xz) - uPillarWidth;
        float k = 4.0;
        float h2 = max(k - abs(d - bound), 0.0);
        d = max(d, bound) + h2*h2*0.0625/k;
        d = abs(d)*0.15 + 0.01;

        float grad = clamp((15.0 - p.y) / 30.0, 0.0, 1.0);
        col += mix(uBottomColor, uTopColor, grad) / d;

        t += d * STEP_MULT;
        if(t > 50.0) break;
      }

      float widthNorm = uPillarWidth / 3.0;
      col = tanh(col * uGlowAmount / widthNorm);
      col -= fract(sin(dot(gl_FragCoord.xy, vec2(12.9898,78.233))) * 43758.5453) / 15.0 * uNoiseIntensity;

      gl_FragColor = vec4(col * uIntensity, 1.0);
    }
  `;

  material = new THREE.ShaderMaterial({
    vertexShader: VS,
    fragmentShader: FS,
    uniforms: {
      uTime:         { value: 0 },
      uResolution:   { value: new THREE.Vector2(w, h) },
      uTopColor:     { value: parseCol(TOP_COLOR) },
      uBottomColor:  { value: parseCol(BOTTOM_COLOR) },
      uIntensity:    { value: INTENSITY },
      uGlowAmount:   { value: GLOW },
      uPillarWidth:  { value: WIDTH },
      uPillarHeight: { value: HEIGHT },
      uNoiseIntensity:{ value: NOISE },
      uRotCos:        { value: 1.0 },
      uRotSin:        { value: 0.0 },
      uPillarRotCos:  { value: Math.cos(pillarRotRad) },
      uPillarRotSin:  { value: Math.sin(pillarRotRad) },
      uWaveSin:       { value: waveSin },
      uWaveCos:       { value: waveCos }
    },
    transparent: true,
    depthWrite: false,
    depthTest:  false
  });

  geometry = new THREE.PlaneGeometry(2, 2);
  scene.add(new THREE.Mesh(geometry, material));

  let lastT = performance.now();
  function frame(now){
    const delta = now - lastT;
    if(delta >= 16){
      timeVal += 0.016 * ROTATION_SPD;
      material.uniforms.uTime.value    = timeVal;
      material.uniforms.uRotCos.value  = Math.cos(timeVal * 0.3);
      material.uniforms.uRotSin.value  = Math.sin(timeVal * 0.3);
      renderer.render(scene, camera);
      lastT = now - (delta % 16);
    }
    rafId = requestAnimationFrame(frame);
  }
  rafId = requestAnimationFrame(frame);

  // Resize
  let resizeT = null;
  window.addEventListener('resize', () => {
    clearTimeout(resizeT);
    resizeT = setTimeout(() => {
      const nw = container.clientWidth  || window.innerWidth;
      const nh = container.clientHeight || window.innerHeight;
      renderer.setSize(nw, nh);
      material.uniforms.uResolution.value.set(nw, nh);
    }, 150);
  }, {passive:true});
})();

// ────────────────────────────────────────────────
// STAGGERED MENU — port vanilla JS de ReactBits
// ────────────────────────────────────────────────
(function(){
  const POSITION = 'left';
  const OFFSCREEN = -100;

  const panel      = document.getElementById('staggered-menu-panel');
  const prelayers  = Array.from(document.querySelectorAll('.sm-prelayer'));
  const toggleBtn  = document.getElementById('sm-toggle-btn');
  const iconEl     = document.getElementById('sm-icon');
  const plusH      = document.getElementById('sm-plus-h');
  const plusV      = document.getElementById('sm-plus-v');
  const textInner  = document.getElementById('sm-text-inner');

  if(!panel || !gsap) return;

  let smOpen = false;
  let busy   = false;

  // Estado inicial
  gsap.set([panel, ...prelayers], { xPercent: OFFSCREEN });
  gsap.set(plusH, { transformOrigin:'50% 50%', rotate:0 });
  gsap.set(plusV, { transformOrigin:'50% 50%', rotate:90 });
  gsap.set(iconEl, { rotate:0, transformOrigin:'50% 50%' });

  function smSetTextLines(lines){
    textInner.innerHTML = lines.map(l=>`<span class="sm-toggle-line">${l}</span>`).join('');
  }

  function smPlayOpen(){
    const itemEls    = Array.from(panel.querySelectorAll('.sm-panel-itemLabel'));
    const numberEls  = Array.from(panel.querySelectorAll('.sm-panel-list[data-numbering] .sm-panel-item'));
    const socialTitle= panel.querySelector('.sm-socials-title');
    const socialLinks= Array.from(panel.querySelectorAll('.sm-socials-link'));

    gsap.set(itemEls,   { yPercent:140, rotate:10 });
    gsap.set(numberEls, { '--sm-num-opacity':0 });
    if(socialTitle) gsap.set(socialTitle, { opacity:0 });
    if(socialLinks.length) gsap.set(socialLinks, { y:25, opacity:0 });

    const tl = gsap.timeline({ onComplete:()=>{ busy=false; } });

    prelayers.forEach((el,i)=>{
      tl.fromTo(el, { xPercent:OFFSCREEN }, { xPercent:0, duration:0.5, ease:'power4.out' }, i*0.07);
    });
    const panelStart = (prelayers.length-1)*0.07 + 0.08;
    tl.fromTo(panel, { xPercent:OFFSCREEN }, { xPercent:0, duration:0.65, ease:'power4.out' }, panelStart);

    if(itemEls.length){
      tl.to(itemEls, { yPercent:0, rotate:0, duration:1, ease:'power4.out', stagger:{ each:0.1, from:'start' } }, panelStart+0.1);
      if(numberEls.length)
        tl.to(numberEls, { '--sm-num-opacity':1, duration:0.6, ease:'power2.out', stagger:{ each:0.08, from:'start' } }, panelStart+0.18);
    }
    if(socialTitle) tl.to(socialTitle, { opacity:1, duration:0.5, ease:'power2.out' }, panelStart+0.26);
    if(socialLinks.length) tl.to(socialLinks, { y:0, opacity:1, duration:0.55, ease:'power3.out', stagger:{ each:0.08, from:'start' } }, panelStart+0.3);

    return tl;
  }

  function smPlayClose(cb){
    gsap.to([...prelayers, panel], {
      xPercent: OFFSCREEN, duration:0.32, ease:'power3.in', overwrite:'auto',
      onComplete(){ if(cb) cb(); busy=false; }
    });
  }

  function smAnimateIcon(opening){
    gsap.to(iconEl, { rotate: opening?225:0, duration: opening?0.8:0.35, ease: opening?'power4.out':'power3.inOut', overwrite:'auto' });
  }

  function smAnimateText(opening){
    const lines = opening ? ['Menú','Cerrar','Menú','Cerrar'] : ['Cerrar','Menú','Cerrar','Menú'];
    const target = opening ? 'Cerrar' : 'Menú';
    smSetTextLines([...lines, target]);
    gsap.set(textInner, { yPercent:0 });
    const total = lines.length + 1;
    gsap.to(textInner, { yPercent: -((total-1)/total*100), duration:0.5+total*0.07, ease:'power4.out' });
  }

  window.smClose = function(){
    if(!smOpen) return;
    smOpen=false;
    toggleBtn.setAttribute('aria-expanded','false');
    panel.setAttribute('aria-hidden','true');
    smPlayClose();
    smAnimateIcon(false);
    smAnimateText(false);
  };

  function smToggle(){
    if(busy) return;
    busy = true;
    smOpen = !smOpen;
    toggleBtn.setAttribute('aria-expanded', smOpen?'true':'false');
    panel.setAttribute('aria-hidden', smOpen?'false':'true');
    if(smOpen){
      smPlayOpen();
    } else {
      smPlayClose();
    }
    smAnimateIcon(smOpen);
    smAnimateText(smOpen);
  }

  toggleBtn.addEventListener('click', smToggle);

  // Cerrar al hacer click fuera
  document.addEventListener('mousedown', e=>{
    if(smOpen && !panel.contains(e.target) && !toggleBtn.contains(e.target)){
      smClose();
    }
  });

  // Nav buttons dentro del panel
  panel.querySelectorAll('.nav-btn[data-page]').forEach(btn=>{
    btn.addEventListener('click', ()=>{
      showPage(btn.dataset.page);
      smClose();
    });
  });

  // Ocultar menú al iniciar (se mostrará solo para admin vía initApp)
  window._smHide = ()=>{ document.getElementById('sm-toggle-btn').style.display='none'; };
  window._smShow = ()=>{ document.getElementById('sm-toggle-btn').style.display='flex'; };
  _smHide();
})();

// ────────────────────────────────────────────────
// ESTADO GLOBAL
// ────────────────────────────────────────────────
let ROL=null,cam1Activa=false,cam2Activa=false,periodoActual="Todo",fechaDesde="",fechaHasta="",statsTab="general",pollingInterval=null,empPollingInterval=null,operarioPollingInterval=null;

// Chart.js instances
let _charts = { barras: null, donut: null, resumen: null };

function _destroyCharts(){
  Object.values(_charts).forEach(c => { if(c) c.destroy(); });
  _charts = { barras: null, donut: null, resumen: null };
}

// Chart.js global defaults (dark theme)
Chart.defaults.color = '#7a7a9a';
Chart.defaults.borderColor = '#252535';
Chart.defaults.font.family = "'JetBrains Mono', monospace";

// ────────────────────────────────────────────────
// HELPERS
// ────────────────────────────────────────────────
function toast(msg,tipo="ok"){
  const el=document.getElementById("toast");
  el.textContent=msg;el.className="show "+tipo;
  setTimeout(()=>{el.className=""},3000);
}
function clock(){
  const n=new Date();
  document.getElementById("hdr-clock").textContent=n.toLocaleDateString("es-CO")+"  "+n.toLocaleTimeString("es-CO");
}
setInterval(clock,1000);clock();
async function api(url,opts={}){
  const res=await fetch(url,{headers:{"Content-Type":"application/json"},...opts});
  if(!res.ok){const j=await res.json().catch(()=>({}));throw new Error(j.error||`HTTP ${res.status}`);}
  return res.json();
}
function onFechaChange(){
  const d=document.getElementById("fecha-desde").value;
  const h=document.getElementById("fecha-hasta").value;
  const lbl=document.getElementById("fecha-label");
  if(d||h)lbl.textContent=`${d||"…"} → ${h||"…"}`;
  else lbl.textContent="";
}
function aplicarFiltroFecha(){
  fechaDesde=document.getElementById("fecha-desde").value;
  fechaHasta=document.getElementById("fecha-hasta").value;
  if(!fechaDesde&&!fechaHasta){alert("Selecciona al menos una fecha.");return;}
  document.querySelectorAll("[data-periodo]").forEach(b=>b.classList.remove("active"));
  document.getElementById("fecha-label").textContent=`✓ Mostrando: ${fechaDesde||"…"} → ${fechaHasta||"…"}`;
  if(statsTab==="general")loadEstadisticas();
  else if(statsTab==="empleado"){const q=document.getElementById("emp-search-input").value.trim();if(q)loadEstadisticas(q);}
}
function limpiarFiltroFecha(){
  fechaDesde="";fechaHasta="";
  document.getElementById("fecha-desde").value="";
  document.getElementById("fecha-hasta").value="";
  document.getElementById("fecha-label").textContent="";
  document.querySelectorAll("[data-periodo]").forEach(b=>{b.classList.toggle("active",b.dataset.periodo===periodoActual);});
  if(statsTab==="general")loadEstadisticas();
  else if(statsTab==="empleado"){const q=document.getElementById("emp-search-input").value.trim();if(q)loadEstadisticas(q);}
}
function _buildStatsUrl(base,empleadoQuery=""){
  let url=`${base}?periodo=${encodeURIComponent(periodoActual)}`;
  if(fechaDesde)url+=`&fecha_desde=${encodeURIComponent(fechaDesde)}`;
  if(fechaHasta)url+=`&fecha_hasta=${encodeURIComponent(fechaHasta)}`;
  if(empleadoQuery)url+=`&empleado=${encodeURIComponent(empleadoQuery)}`;
  return url;
}

// ────────────────────────────────────────────────
// POLLING MODELOS
// ────────────────────────────────────────────────
(async function pollModelos(){
  try{
    const d=await fetch("/api/modelos_estado").then(r=>r.json());
    const dot=document.getElementById("modelos-dot");
    const txt=document.getElementById("modelos-txt");
    const box=document.getElementById("modelos-status");
    if(d.listos){
      dot.style.animation="none";dot.style.background="var(--green)";
      txt.style.color="var(--green)";txt.textContent="✓ Modelos listos — detección de IA activa";
      box.style.borderColor="rgba(0,230,118,.25)";box.style.background="rgba(0,230,118,.07)";
      const banner=document.getElementById("banner-modelos");
      if(banner)banner.style.display="none";
    }else{setTimeout(pollModelos,2000);}
  }catch{setTimeout(pollModelos,3000);}
})();

// ────────────────────────────────────────────────
// LOGIN / LOGOUT
// ────────────────────────────────────────────────
document.getElementById("l-pass").addEventListener("keydown",e=>{if(e.key==="Enter")doLogin();});
document.getElementById("l-user").addEventListener("keydown",e=>{if(e.key==="Enter")document.getElementById("l-pass").focus();});
document.getElementById("btn-login").onclick=doLogin;
async function doLogin(){
  const usuario=document.getElementById("l-user").value.trim();
  const password=document.getElementById("l-pass").value.trim();
  document.getElementById("l-err").textContent="";
  if(!usuario||!password){document.getElementById("l-err").textContent="✗ Ingresa usuario y contraseña";return;}
  try{
    const data=await api("/api/login",{method:"POST",body:JSON.stringify({usuario,password})});
    if(data.ok){
      ROL=data.rol;
      initApp(ROL);
      if(!data.modelos_listos){
        document.getElementById("banner-modelos").style.display="flex";
      }
    }else{
      document.getElementById("l-err").textContent="✗ "+(data.error||"Error al ingresar");
    }
  }catch(e){
    document.getElementById("l-err").textContent="✗ Error de conexión";
  }
}
document.getElementById("btn-logout").onclick=async()=>{
  cam1Activa=false; cam2Activa=false;
  await fetch("/api/camara/detener",{method:"POST"}).catch(()=>{});
  await fetch("/api/logout",{method:"POST"}).catch(()=>{});
  resetUI();
  document.getElementById("login-screen").style.display="flex";
  document.getElementById("l-user").value="";
  document.getElementById("l-pass").value="";
  document.getElementById("l-err").textContent="";
  ROL=null;
};

// ────────────────────────────────────────────────
// INIT APP
// ────────────────────────────────────────────────
function resetUI(){
  // Cierra el menú staggered si está abierto
  if(typeof smClose==='function') smClose();
  // Limpia TODO antes de cada sesión
  clearInterval(pollingInterval);
  clearInterval(empPollingInterval);
  clearInterval(operarioPollingInterval);
  pollingInterval=null; empPollingInterval=null; operarioPollingInterval=null;
  _destroyCharts();
  document.querySelectorAll(".page").forEach(p=>{ p.classList.remove("active"); });
  document.querySelectorAll(".nav-btn").forEach(b=>b.classList.remove("active"));
  // Ocultar todo
  document.getElementById("app").classList.remove("visible");
  document.getElementById("sidebar").style.cssText="";
  document.getElementById("main-content").style.cssText="";
  document.getElementById("page-empleado").style.cssText="display:none";
  document.getElementById("banner-modelos").style.display="none";
}

function initApp(rol){
  resetUI();
  document.getElementById("login-screen").style.display="none";
  document.getElementById("app").classList.add("visible");
  if(rol==="administrador"){
    _smShow();
    document.getElementById("sidebar").style.display="none";
    document.getElementById("main-content").style.display="block";
    document.getElementById("page-empleado").style.display="none";
    showPage("camara");
    startPolling();
    loadConfig();
  } else {
    _smHide();
    document.getElementById("sidebar").style.display="none";
    document.getElementById("main-content").style.display="none";
    document.getElementById("page-empleado").style.cssText="display:flex;flex:1;overflow-y:auto";
    startEmpPolling();
  }
}

// ────────────────────────────────────────────────
// NAVEGACIÓN
// ────────────────────────────────────────────────
document.querySelectorAll(".nav-btn[data-page]").forEach(btn=>{
  btn.addEventListener("click",()=>showPage(btn.dataset.page));
});
function showPage(name){
  document.querySelectorAll(".page").forEach(p=>p.classList.remove("active"));
  document.querySelectorAll(".nav-btn").forEach(b=>b.classList.remove("active"));
  const pg=document.getElementById("page-"+name);if(pg)pg.classList.add("active");
  const btn=document.querySelector(`.nav-btn[data-page="${name}"]`);if(btn)btn.classList.add("active");
  if(name==="registros")loadRegistros();
  if(name==="eventos")loadEventos();
  if(name==="estadisticas")loadEstadisticas();
  if(name==="evidencias")cargarEvidencias();
}

// ────────────────────────────────────────────────
// POLLING ESTADO CÁMARA
// ────────────────────────────────────────────────
function startPolling(){
  clearInterval(pollingInterval);
  clearInterval(operarioPollingInterval);
  pollingInterval=setInterval(pollEstado,800);
  pollOperario();
  operarioPollingInterval=setInterval(pollOperario,3000);
}
async function pollEstado(){try{const d=await api("/api/camara/estado");updateCamUI(d);}catch{}}
async function pollOperario(){
  try{
    const d=await api("/api/operario");
    const ops=d.operarios||[];
    if(ops.length>0){
      const textos=ops.map(o=>o.nombre.toUpperCase()).join(" / ");
      const subs=ops.map(o=>o.cargo+" · Entrada: "+o.entrada).join(" | ");
      document.getElementById("sb-op-name").textContent=textos;
      document.getElementById("sb-op-name").className="sidebar-op-name";
      document.getElementById("sb-op-sub").textContent=subs;
    }else{
      document.getElementById("sb-op-name").textContent="— Sin registro —";
      document.getElementById("sb-op-name").className="sidebar-op-name none";
      document.getElementById("sb-op-sub").textContent="";
    }
    // Actualizar labels de cámaras admin
    const lbl1=document.getElementById("admin-op1-label");
    const lbl2=document.getElementById("admin-op2-label");
    if(lbl1){
      const op1=ops.find(o=>o.slot===0);
      lbl1.textContent=op1?"● "+op1.nombre.toUpperCase():"";
    }
    if(lbl2){
      const op2=ops.find(o=>o.slot===1);
      lbl2.textContent=op2?"● "+op2.nombre.toUpperCase():"";
    }
  }catch{}
}
function updateCamUI(d){
  const dot=document.getElementById("cam-dot");
  const badgeTxt=document.getElementById("cam-badge-txt");
  if(d.running){dot.classList.add("active");badgeTxt.textContent=d.running2?"CÁMARAS ACTIVAS":"CÁMARA 1 ACTIVA";}
  else{dot.classList.remove("active");badgeTxt.textContent=d.running2?"CÁMARA 2 ACTIVA":"SIN CÁMARA";}
  document.getElementById("dot1").className="dot"+(d.running?" on":"");
  document.getElementById("dot2").className="dot"+(d.running2?" on":"");
  // ── Cam1 status ──
  const ojos=d.ojos||"SIN DETECCIÓN";
  const oel=document.getElementById("st-ojos");
  oel.textContent=ojos;oel.className="s-val "+(ojos.includes("FATIGA")?"s-warn":ojos==="ACTIVO"?"s-ok":"");
  document.getElementById("st-ear").textContent=d.ear?.toFixed(2)??"0.00";
  const pel=document.getElementById("st-persona");
  pel.textContent=d.persona?"DETECTADA":"NO DETECTADA";pel.className="s-val "+(d.persona?"s-ok":"");
  const cel=document.getElementById("st-casco");
  cel.textContent=d.epp_ok?.includes("CASCO")?"[✓] PRESENTE":"[✗] FALTANTE";cel.className="s-val "+(d.epp_ok?.includes("CASCO")?"s-ok":"s-warn");
  const chal=document.getElementById("st-chaleco");
  chal.textContent=d.epp_ok?.includes("CHALECO")?"[✓] PRESENTE":"[✗] FALTANTE";chal.className="s-val "+(d.epp_ok?.includes("CHALECO")?"s-ok":"s-warn");
  const tel=document.getElementById("st-tel");
  tel.textContent=d.telefono?"!! EN USO !!":"SIN USO";tel.className="s-val "+(d.telefono?"s-warn":"s-ok");
  // ── Cam2 status ──
  const ojos2=d.ojos2||"SIN DETECCIÓN";
  const oel2=document.getElementById("st-ojos2");
  if(oel2){oel2.textContent=d.running2?ojos2:"—";oel2.className="s-val "+(ojos2.includes("FATIGA")?"s-warn":ojos2==="ACTIVO"?"s-ok":"");}
  const ear2el=document.getElementById("st-ear2");
  if(ear2el)ear2el.textContent=d.running2?(d.ear2?.toFixed(2)??"0.00"):"—";
  const pel2=document.getElementById("st-persona2");
  if(pel2){pel2.textContent=d.running2?(d.persona2?"DETECTADA":"NO DETECTADA"):"—";pel2.className="s-val "+(d.persona2?"s-ok":"");}
  const cel2=document.getElementById("st-casco2");
  if(cel2){cel2.textContent=d.running2?(d.epp_ok2?.includes("CASCO")?"[✓] PRESENTE":"[✗] FALTANTE"):"—";cel2.className="s-val "+(d.epp_ok2?.includes("CASCO")?"s-ok":"s-warn");}
  const chal2=document.getElementById("st-chaleco2");
  if(chal2){chal2.textContent=d.running2?(d.epp_ok2?.includes("CHALECO")?"[✓] PRESENTE":"[✗] FALTANTE"):"—";chal2.className="s-val "+(d.epp_ok2?.includes("CHALECO")?"s-ok":"s-warn");}
  const tel2=document.getElementById("st-tel2");
  if(tel2){tel2.textContent=d.running2?(d.telefono2?"!! EN USO !!":"SIN USO"):"—";tel2.className="s-val "+(d.telefono2?"s-warn":"s-ok");}
  // ── Alertas combinadas ──
  const bar=document.getElementById("alerts-bar");
  bar.innerHTML="";
  if(!d.alertas?.length){bar.innerHTML='<div class="no-alerts">✓ Sin alertas activas</div>';}
  else{d.alertas.forEach(a=>{const cls=a.includes("FATIGA")?"red":a.includes("TELÉFONO")?"yellow":"orange";bar.innerHTML+=`<div class="alert-chip ${cls}">⚠ ${a.replace("⚠ ","")}</div>`;});}
}

// ────────────────────────────────────────────────
// CÁMARAS TOGGLE (Admin — botones independientes)
// ────────────────────────────────────────────────
async function toggleCam(num){
  if(num===1){
    const btn=document.getElementById("btn-toggle-cam1");
    if(!cam1Activa){
      try{
        await api("/api/camara/iniciar/1",{method:"POST"});
        cam1Activa=true;
        btn.textContent="■ APAGAR CÁMARA 1";btn.style.background="var(--red)";
        document.getElementById("feed1").src="/video/cam1?"+Date.now();
        document.getElementById("feed1").style.display="block";
        document.getElementById("placeholder1").style.display="none";
        toast("Cámara 1 iniciada","ok");
      }catch(e){toast("✗ Error cam1: "+e.message,"err");}
    }else{
      try{
        await api("/api/camara/detener/1",{method:"POST"});
        cam1Activa=false;
        btn.textContent="▶ ENCENDER CÁMARA 1";btn.style.background="var(--accent)";
        document.getElementById("feed1").src="";
        document.getElementById("feed1").style.display="none";
        document.getElementById("placeholder1").style.display="flex";
        toast("Cámara 1 detenida","ok");
      }catch(e){toast("✗ Error cam1: "+e.message,"err");}
    }
  } else {
    const btn=document.getElementById("btn-toggle-cam2");
    if(!cam2Activa){
      try{
        await api("/api/camara/iniciar/2",{method:"POST"});
        cam2Activa=true;
        btn.textContent="■ APAGAR CÁMARA 2";btn.style.background="var(--red)";
        document.getElementById("feed2").src="/video/cam2?"+Date.now();
        document.getElementById("feed2").style.display="block";
        document.getElementById("placeholder2").style.display="none";
        toast("Cámara 2 iniciada","ok");
      }catch(e){toast("✗ Error cam2: "+e.message,"err");}
    }else{
      try{
        await api("/api/camara/detener/2",{method:"POST"});
        cam2Activa=false;
        btn.textContent="▶ ENCENDER CÁMARA 2";btn.style.background="var(--blue)";
        document.getElementById("feed2").src="";
        document.getElementById("feed2").style.display="none";
        document.getElementById("placeholder2").style.display="flex";
        toast("Cámara 2 detenida","ok");
      }catch(e){toast("✗ Error cam2: "+e.message,"err");}
    }
  }
}

// ────────────────────────────────────────────────
// MODALES
// ────────────────────────────────────────────────
function openModal(tipo){
  document.getElementById("modal-"+tipo).classList.add("open");
  if(tipo==="entrada"){
    document.getElementById("m-ced").value="";document.getElementById("m-nom").value="";
    document.getElementById("m-car").value="";document.getElementById("m-ent-err").textContent="";
    document.getElementById("m-ced").focus();
  }else{document.getElementById("m-sal-ced").value="";document.getElementById("m-sal-err").textContent="";document.getElementById("m-sal-ced").focus();}
}
function closeModal(tipo){document.getElementById("modal-"+tipo).classList.remove("open");}
document.querySelectorAll(".modal-overlay").forEach(el=>{el.addEventListener("click",e=>{if(e.target===el)el.classList.remove("open");});});
async function confirmarEntrada(){
  const cedula=document.getElementById("m-ced").value.trim();
  const nombre=document.getElementById("m-nom").value.trim();
  const cargo=document.getElementById("m-car").value.trim();
  const errEl=document.getElementById("m-ent-err");
  if(!cedula||!nombre||!cargo){errEl.textContent="Por favor completa todos los campos.";return;}
  try{
    const d=await api("/api/entrada",{method:"POST",body:JSON.stringify({cedula,nombre,cargo})});
    if(d.ok){
      closeModal("entrada");
      toast("✓ Entrada registrada — "+nombre,"ok");
      const slot=d.slot; // 0 o 1
      // Si es portal empleado, mostrar datos y encender la cámara correspondiente
      if(ROL==="empleado"){
        if(slot===0){
          document.getElementById("emp-op1-status").textContent=nombre.toUpperCase()+" · "+cargo;
          document.getElementById("emp-op1-status").className="emp-status active";
          try{
            const r=await api("/api/camara/iniciar/1",{method:"POST"});
            if(r.ok||r.msg){
              // Pequeño delay para que el stream arranque
              setTimeout(()=>_syncEmpCamFeed(0,true),500);
            }
          }catch(ex){toast("⚠ Cámara 1 no disponible: "+ex.message,"err");}
        } else {
          document.getElementById("emp-card-op2").style.display="block";
          document.getElementById("emp-op2-status").textContent=nombre.toUpperCase()+" · "+cargo;
          document.getElementById("emp-op2-status").className="emp-status active";
          try{
            const r=await api("/api/camara/iniciar/2",{method:"POST"});
            if(r.ok||r.msg){
              setTimeout(()=>_syncEmpCamFeed(1,true),500);
            }
          }catch(ex){toast("⚠ Cámara 2 no disponible: "+ex.message,"err");}
        }
      }
    }
    else errEl.textContent=d.error;
  }catch(e){errEl.textContent=e.message;}
}
async function confirmarSalida(){
  const cedula=document.getElementById("m-sal-ced").value.trim();
  const errEl=document.getElementById("m-sal-err");
  if(!cedula){errEl.textContent="Ingresa la cédula.";return;}
  try{
    const d=await api("/api/salida",{method:"POST",body:JSON.stringify({cedula})});
    if(d.ok){
      closeModal("salida");
      toast("✓ Salida registrada","ok");
      const slot=d.slot;
      // Si es portal empleado, actualizar solo el slot que salió
      if(ROL==="empleado"){
        if(slot===0){
          document.getElementById("emp-op1-status").textContent="— Ninguno registrado —";
          document.getElementById("emp-op1-status").className="emp-status none";
          _syncEmpCamFeed(0, false);
          // No tocar la cámara 2 ni el operario 2
        } else {
          document.getElementById("emp-op2-status").textContent="— Ninguno registrado —";
          document.getElementById("emp-op2-status").className="emp-status none";
          _syncEmpCamFeed(1, false);
          document.getElementById("emp-card-op2").style.display="none";
          // No tocar la cámara 1 ni el operario 1
        }
      }
    }
    else errEl.textContent=d.error;
  }catch(e){errEl.textContent=e.message;}
}

// ────────────────────────────────────────────────
// REGISTROS
// ────────────────────────────────────────────────
async function loadRegistros(){
  const rows=await api("/api/registros");
  const tbody=document.getElementById("tbody-registros");
  if(!rows.length){tbody.innerHTML='<tr><td colspan="6" style="text-align:center;color:var(--sub);padding:24px">Sin registros aún.</td></tr>';return;}
  tbody.innerHTML=rows.map(r=>`<tr><td>${r.Cedula||""}</td><td>${r.Nombre||""}</td><td>${r.Cargo||""}</td><td>${r.Entrada||""}</td><td>${r.Salida||"—"}</td><td>${r.Duracion_min||"—"}</td></tr>`).join("");
}

// ────────────────────────────────────────────────
// EVENTOS
// ────────────────────────────────────────────────
async function loadEventos(){
  const rows=await api("/api/eventos");
  const tbody=document.getElementById("tbody-eventos");
  if(!rows.length){tbody.innerHTML='<tr><td colspan="3" style="text-align:center;color:var(--sub);padding:24px">Sin eventos registrados.</td></tr>';return;}
  tbody.innerHTML=rows.map(ev=>{
    const tipo=ev.tipo||"";
    let tag="gray";
    if(tipo.includes("FATIGA"))tag="red";
    else if(tipo.includes("TELEFONO"))tag="yellow";
    else if(tipo.includes("EPP"))tag="orange";
    else if(tipo.includes("ENTRADA"))tag="green";
    else if(tipo.includes("SALIDA"))tag="blue";
    return `<tr><td>${ev.ts}</td><td>${ev.cedula}</td><td><span class="tag tag-${tag}">${tipo}</span></td></tr>`;
  }).join("");
}

// ────────────────────────────────────────────────
// ESTADÍSTICAS — Chart.js
// ────────────────────────────────────────────────
document.querySelectorAll("[data-periodo]").forEach(btn=>{
  btn.addEventListener("click",()=>{
    document.querySelectorAll("[data-periodo]").forEach(b=>b.classList.remove("active"));
    btn.classList.add("active");periodoActual=btn.dataset.periodo;
    if(statsTab==="general")loadEstadisticas();
  });
});
function switchStatsTab(tab){
  statsTab=tab;
  document.getElementById("stats-general").style.display=tab==="general"?"":"none";
  document.getElementById("stats-empleado").style.display=tab==="empleado"?"":"none";
  document.getElementById("tab-general-btn").classList.toggle("active",tab==="general");
  document.getElementById("tab-emp-btn").classList.toggle("active",tab==="empleado");
  if(tab==="general")loadEstadisticas();
}

function _renderCharts(d){
  _destroyCharts();
  const c=d.conteo||{};
  const e=d.epp_detalle||{};
  const chartsRow=document.getElementById("charts-row");
  chartsRow.innerHTML=`
    <div class="chart-card chart-full">
      <div class="chart-card-title">📊 Eventos por tipo — ${d.periodo||""}</div>
      <div class="chart-container" style="height:240px"><canvas id="ch-barras"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">🛡 EPP Faltante</div>
      <div class="chart-container" style="height:220px"><canvas id="ch-donut"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">⚡ Resumen de Incidencias</div>
      <div class="chart-container" style="height:220px"><canvas id="ch-resumen"></canvas></div>
    </div>`;

  // Chart 1 — Barras de eventos
  const eventLabels=Object.keys(c);
  const eventVals=Object.values(c);
  const barColors=eventLabels.map(k=>{
    if(k.includes("FATIGA"))return "rgba(255,61,90,.8)";
    if(k.includes("TELEFONO"))return "rgba(68,138,255,.8)";
    if(k.includes("EPP"))return "rgba(255,107,0,.8)";
    return "rgba(122,122,154,.8)";
  });
  if(eventLabels.length){
    _charts.barras=new Chart(document.getElementById("ch-barras"),{
      type:"bar",
      data:{labels:eventLabels,datasets:[{
        label:"Eventos",data:eventVals,
        backgroundColor:barColors,
        borderColor:barColors.map(c=>c.replace(".8","1")),
        borderWidth:1,borderRadius:6,borderSkipped:false
      }]},
      options:{
        responsive:true,maintainAspectRatio:false,
        plugins:{
          legend:{display:false},
          tooltip:{callbacks:{label:ctx=>" "+ctx.raw+" eventos"}}
        },
        scales:{
          x:{grid:{color:"#252535"},ticks:{color:"#7a7a9a"}},
          y:{grid:{color:"#252535"},ticks:{color:"#7a7a9a",stepSize:1},beginAtZero:true}
        }
      }
    });
  }else{
    document.getElementById("ch-barras").parentElement.innerHTML='<div class="chart-empty">Sin eventos registrados</div>';
  }

  // Chart 2 — Donut EPP
  const eppLabels=Object.keys(e);
  const eppVals=Object.values(e);
  if(eppLabels.length){
    _charts.donut=new Chart(document.getElementById("ch-donut"),{
      type:"doughnut",
      data:{labels:eppLabels,datasets:[{
        data:eppVals,
        backgroundColor:["rgba(255,107,0,.85)","rgba(255,202,40,.85)","rgba(255,61,90,.85)","rgba(68,138,255,.85)"],
        borderColor:"#16161f",borderWidth:3,hoverOffset:8
      }]},
      options:{
        responsive:true,maintainAspectRatio:false,cutout:"65%",
        plugins:{
          legend:{position:"bottom",labels:{color:"#7a7a9a",padding:12,font:{size:11}}},
          tooltip:{callbacks:{label:ctx=>` ${ctx.label}: ${ctx.raw} eventos`}}
        }
      }
    });
  }else{
    document.getElementById("ch-donut").parentElement.innerHTML='<div class="chart-empty">Sin datos de EPP</div>';
  }

  // Chart 3 — Barras horizontales resumen
  const fatiga=c.FATIGA||0;
  const telefono=c.TELEFONO||0;
  const eppTot=Object.values(e).reduce((a,b)=>a+b,0);
  _charts.resumen=new Chart(document.getElementById("ch-resumen"),{
    type:"bar",
    data:{
      labels:["Fatiga","Teléfono","EPP Faltante"],
      datasets:[{
        label:"Incidencias",
        data:[fatiga,telefono,eppTot],
        backgroundColor:["rgba(255,61,90,.8)","rgba(68,138,255,.8)","rgba(255,107,0,.8)"],
        borderColor:["#ff3d5a","#448aff","#ff6b00"],
        borderWidth:1,borderRadius:6
      }]
    },
    options:{
      indexAxis:"y",
      responsive:true,maintainAspectRatio:false,
      plugins:{
        legend:{display:false},
        tooltip:{callbacks:{label:ctx=>` ${ctx.raw} incidencias`}}
      },
      scales:{
        x:{grid:{color:"#252535"},ticks:{color:"#7a7a9a",stepSize:1},beginAtZero:true},
        y:{grid:{color:"#252535"},ticks:{color:"#7a7a9a"}}
      }
    }
  });
}

async function loadEstadisticas(empleadoQuery=""){
  const kpiRow=document.getElementById("kpi-row");
  const chartsRow=document.getElementById("charts-row");
  kpiRow.innerHTML='<div style="grid-column:1/-1;color:var(--sub);font-size:12px;padding:10px">Cargando indicadores...</div>';
  chartsRow.innerHTML='<div class="chart-card chart-full"><div class="chart-empty">Cargando gráficas...</div></div><div class="chart-card"><div class="chart-empty">Cargando...</div></div><div class="chart-card"><div class="chart-empty">Cargando...</div></div>';
  let url=_buildStatsUrl("/api/estadisticas",empleadoQuery);
  try{
    const d=await fetch(url).then(r=>r.json());
    if(d.error){kpiRow.innerHTML=`<div style="color:var(--red);grid-column:1/-1;padding:10px">${d.error}</div>`;return;}
    const c=d.conteo||{};const e=d.epp_detalle||{};
    const kpis=[
      {lbl:"Turnos",val:d.n_turnos,col:"var(--blue)",icon:"⏱"},
      {lbl:"Fatiga",val:c.FATIGA||0,col:"var(--red)",icon:"😴"},
      {lbl:"Casco faltante",val:e.CASCO||0,col:"var(--accent)",icon:"⛑"},
      {lbl:"Chaleco faltante",val:e.CHALECO||0,col:"var(--yellow)",icon:"🦺"},
      {lbl:"Uso teléfono",val:c.TELEFONO||0,col:"var(--blue)",icon:"📱"},
    ];
    kpiRow.innerHTML=kpis.map(k=>`<div class="kpi-card"><div class="kpi-card-bar" style="background:${k.col}"></div><div style="font-size:18px;margin-bottom:4px">${k.icon}</div><div class="kpi-val" style="color:${k.col}">${k.val}</div><div class="kpi-lbl">${k.lbl}</div></div>`).join("");
    _renderCharts(d);
  }catch(err){kpiRow.innerHTML=`<div style="color:var(--red);grid-column:1/-1;padding:10px">Error: ${err.message}</div>`;}
}

async function buscarEmpleado(){
  const query=document.getElementById("emp-search-input").value.trim();
  document.getElementById("emp-suggestions").style.display="none";
  if(!query)return;
  const res=document.getElementById("emp-results");
  res.innerHTML='<div style="color:var(--sub);font-size:12px;padding:24px">Buscando...</div>';
  let url=_buildStatsUrl("/api/estadisticas",query);
  try{
    const d=await fetch(url).then(r=>r.json());
    if(d.error){res.innerHTML=`<div style="color:var(--red);padding:24px">${d.error}</div>`;return;}
    const c=d.conteo||{};const e=d.epp_detalle||{};
    const dur=d.duraciones||[];
    const totalMin=dur.reduce((a,b)=>a+b,0);
    const horas=(totalMin/60).toFixed(1);
    const nombreEmp=d.nombre_emp||query;
    const cedulaEmp=d.cedula_emp||"";
    const cargoEmp=d.cargo_emp||"";
    const regs=d.registros_emp||[];

    // Encabezado del empleado
    res.innerHTML=`
      <div style="background:rgba(22,22,31,0.88);border:1px solid var(--border);border-radius:10px;padding:14px 18px;margin-bottom:12px">
        <div style="color:var(--accent);font-weight:700;font-size:14px">👤 ${nombreEmp.toUpperCase()}</div>
        <div style="color:var(--sub);font-size:11px;margin-top:4px">Cédula: ${cedulaEmp} &nbsp;•&nbsp; Cargo: ${cargoEmp} &nbsp;•&nbsp; Período: ${d.periodo||""}</div>
      </div>
      <div class="kpi-row" style="grid-template-columns:repeat(6,1fr);margin-bottom:14px" id="emp-kpi-row">
        <div class="kpi-card"><div class="kpi-card-bar" style="background:var(--blue)"></div><div style="font-size:16px;margin-bottom:2px">⏱</div><div class="kpi-val" style="color:var(--blue);font-size:22px">${d.n_turnos}</div><div class="kpi-lbl">Turnos</div></div>
        <div class="kpi-card"><div class="kpi-card-bar" style="background:var(--blue)"></div><div style="font-size:16px;margin-bottom:2px">🕐</div><div class="kpi-val" style="color:var(--blue);font-size:22px">${horas}h</div><div class="kpi-lbl">Horas totales</div></div>
        <div class="kpi-card"><div class="kpi-card-bar" style="background:var(--red)"></div><div style="font-size:16px;margin-bottom:2px">😴</div><div class="kpi-val" style="color:var(--red);font-size:22px">${c.FATIGA||0}</div><div class="kpi-lbl">Fatiga</div></div>
        <div class="kpi-card"><div class="kpi-card-bar" style="background:var(--accent)"></div><div style="font-size:16px;margin-bottom:2px">⛑</div><div class="kpi-val" style="color:var(--accent);font-size:22px">${e.CASCO||0}</div><div class="kpi-lbl">Casco faltante</div></div>
        <div class="kpi-card"><div class="kpi-card-bar" style="background:var(--yellow)"></div><div style="font-size:16px;margin-bottom:2px">🦺</div><div class="kpi-val" style="color:var(--yellow);font-size:22px">${e.CHALECO||0}</div><div class="kpi-lbl">Chaleco faltante</div></div>
        <div class="kpi-card"><div class="kpi-card-bar" style="background:var(--blue)"></div><div style="font-size:16px;margin-bottom:2px">📱</div><div class="kpi-val" style="color:var(--blue);font-size:22px">${c.TELEFONO||0}</div><div class="kpi-lbl">Uso teléfono</div></div>
      </div>
      <div class="charts-grid" id="emp-charts-grid">
        <div class="chart-card"><div class="chart-card-title">Eventos por tipo</div><div class="chart-container" style="height:200px"><canvas id="ch-emp-barras"></canvas></div></div>
        <div class="chart-card"><div class="chart-card-title">EPP Faltante</div><div class="chart-container" style="height:200px"><canvas id="ch-emp-donut"></canvas></div></div>
        ${dur.length?`<div class="chart-card chart-full"><div class="chart-card-title">Duración de Turnos (min)</div><div class="chart-container" style="height:180px"><canvas id="ch-emp-dur"></canvas></div></div>`:''}
      </div>
      ${regs.length?`
      <div style="margin-top:14px">
        <div style="font-size:10px;color:var(--sub);text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px;font-weight:700">Historial de Turnos</div>
        <div style="background:rgba(22,22,31,0.88);border:1px solid var(--border);border-radius:10px;overflow:hidden">
          <table style="width:100%;border-collapse:collapse;font-size:12px">
            <thead><tr>
              <th style="background:var(--border);color:var(--accent);font-weight:700;padding:10px 14px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.06em">Entrada</th>
              <th style="background:var(--border);color:var(--accent);font-weight:700;padding:10px 14px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.06em">Salida</th>
              <th style="background:var(--border);color:var(--accent);font-weight:700;padding:10px 14px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.06em">Duración (min)</th>
            </tr></thead>
            <tbody>${regs.map(r=>`<tr style="border-bottom:1px solid var(--border)"><td style="padding:10px 14px">${r.Entrada||""}</td><td style="padding:10px 14px">${r.Salida||"—"}</td><td style="padding:10px 14px">${r.Duracion_min||"—"}</td></tr>`).join("")}</tbody>
          </table>
        </div>
      </div>`:''}`;

    // Gráfico barras eventos
    const eLabels=Object.keys(c),eVals=Object.values(c);
    if(eLabels.length){
      const bColors=eLabels.map(k=>k.includes("FATIGA")?"rgba(255,61,90,.8)":k.includes("TELEFONO")?"rgba(68,138,255,.8)":k.includes("EPP")?"rgba(255,107,0,.8)":"rgba(122,122,154,.8)");
      new Chart(document.getElementById("ch-emp-barras"),{type:"bar",data:{labels:eLabels,datasets:[{data:eVals,backgroundColor:bColors,borderWidth:1,borderRadius:4}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{x:{grid:{color:"#252535"},ticks:{color:"#7a7a9a"}},y:{grid:{color:"#252535"},ticks:{color:"#7a7a9a",stepSize:1},beginAtZero:true}}}});
    }
    // Gráfico donut EPP
    const dLabels=Object.keys(e),dVals=Object.values(e);
    if(dLabels.length){
      new Chart(document.getElementById("ch-emp-donut"),{type:"doughnut",data:{labels:dLabels,datasets:[{data:dVals,backgroundColor:["rgba(255,107,0,.85)","rgba(255,202,40,.85)"],borderColor:"#16161f",borderWidth:3,hoverOffset:6}]},options:{responsive:true,maintainAspectRatio:false,cutout:"60%",plugins:{legend:{position:"bottom",labels:{color:"#7a7a9a",font:{size:11}}}}}});
    }
    // Gráfico duración de turnos
    if(dur.length){
      const durLabels=dur.map((_,i)=>`Turno ${i+1}`);
      new Chart(document.getElementById("ch-emp-dur"),{type:"bar",data:{labels:durLabels,datasets:[{label:"Minutos",data:dur,backgroundColor:"rgba(68,138,255,.8)",borderColor:"#448aff",borderWidth:1,borderRadius:4}]},options:{indexAxis:"y",responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{x:{grid:{color:"#252535"},ticks:{color:"#7a7a9a"},beginAtZero:true,title:{display:true,text:"Minutos",color:"#7a7a9a"}},y:{grid:{color:"#252535"},ticks:{color:"#7a7a9a"}}}}});
    }
  }catch(err){res.innerHTML=`<div style="color:var(--red);padding:24px">Error: ${err.message}</div>`;}
}

// ────────────────────────────────────────────────
// AUTOCOMPLETADO EMPLEADOS
// ────────────────────────────────────────────────
let _empListCache=null;
async function _getEmpleados(){
  if(_empListCache)return _empListCache;
  try{const d=await api("/api/empleados");_empListCache=d;return d;}catch{return[];}
}
async function filtrarSugerencias(q){
  const sug=document.getElementById("emp-suggestions");
  if(!q||q.length<2){sug.style.display="none";return;}
  const lista=await _getEmpleados();
  const ql=q.toLowerCase();
  const matches=lista.filter(e=>e.nombre.toLowerCase().includes(ql)||e.cedula.toLowerCase().includes(ql)).slice(0,6);
  if(!matches.length){sug.style.display="none";return;}
  sug.innerHTML=matches.map(e=>`<div onclick="seleccionarSugerencia('${e.nombre.replace(/'/g,"\\'")}')\" style="padding:9px 14px;cursor:pointer;font-size:12px;color:var(--text);border-bottom:1px solid var(--border)" onmouseover="this.style.background='var(--border)'" onmouseout="this.style.background=''"><span style="color:var(--accent);font-weight:700">${e.nombre}</span> <span style="color:var(--sub);font-size:11px">(${e.cedula})</span></div>`).join("");
  sug.style.display="block";
}
function seleccionarSugerencia(nombre){
  document.getElementById("emp-search-input").value=nombre;
  document.getElementById("emp-suggestions").style.display="none";
  buscarEmpleado();
}
document.addEventListener("click",e=>{if(!e.target.closest("#emp-suggestions")&&!e.target.closest("#emp-search-input")){const s=document.getElementById("emp-suggestions");if(s)s.style.display="none";}});

function clearEmpSearch(){
  _empListCache=null;
  document.getElementById("emp-search-input").value="";
  const sug=document.getElementById("emp-suggestions");if(sug)sug.style.display="none";
  document.getElementById("emp-results").innerHTML='<div style="color:var(--sub);font-size:12px;text-align:center;padding:40px">Ingresa el nombre o cédula de un operario para ver su desempeño.</div>';
}

// ────────────────────────────────────────────────
// EXPORTAR
// ────────────────────────────────────────────────
function exportarExcel(tipo){window.location=`/api/exportar/${tipo}_excel`;}
async function exportarPDF(){
  const empQuery=statsTab==="empleado"?(document.getElementById("emp-search-input")?.value?.trim()||""):"";
  const url=_buildStatsUrl("/api/exportar/pdf",empQuery);
  toast("Generando PDF...","ok");window.location=url;
}

// ────────────────────────────────────────────────
// CONFIGURACIÓN
// ────────────────────────────────────────────────
async function loadConfig(){
  try{
    const d=await api("/api/config");
    document.getElementById("cfg-cam1").value=d.cam1_index??0;document.getElementById("cfg-cam1-val").textContent=d.cam1_index??0;
    document.getElementById("cfg-cam2").value=d.cam2_index??1;document.getElementById("cfg-cam2-val").textContent=d.cam2_index??1;
    document.getElementById("cfg-conf").value=d.conf_umbral??0.25;document.getElementById("cfg-conf-val").textContent=(d.conf_umbral??0.25).toFixed(2);
    document.getElementById("cfg-ear").value=d.ear_umbral??0.25;document.getElementById("cfg-ear-val").textContent=(d.ear_umbral??0.25).toFixed(2);
    document.getElementById("cfg-tfat").value=d.tiempo_umbral??5;document.getElementById("cfg-tfat-val").textContent=(d.tiempo_umbral??5).toFixed(1);
    document.getElementById("cfg-gracia").value=d.epp_gracia??4;document.getElementById("cfg-gracia-val").textContent=(d.epp_gracia??4).toFixed(1);
    document.getElementById("cfg-casco").checked=d.epp_requerido?.includes("CASCO")??true;
    document.getElementById("cfg-chaleco").checked=d.epp_requerido?.includes("CHALECO")??true;
  }catch{}
}
async function guardarConfig(){
  const epp=[];
  if(document.getElementById("cfg-casco").checked)epp.push("CASCO");
  if(document.getElementById("cfg-chaleco").checked)epp.push("CHALECO");
  try{
    await api("/api/config",{method:"POST",body:JSON.stringify({
      cam1_index:parseInt(document.getElementById("cfg-cam1").value),
      cam2_index:parseInt(document.getElementById("cfg-cam2").value),
      conf_umbral:parseFloat(document.getElementById("cfg-conf").value),
      ear_umbral:parseFloat(document.getElementById("cfg-ear").value),
      tiempo_umbral:parseFloat(document.getElementById("cfg-tfat").value),
      epp_gracia:parseFloat(document.getElementById("cfg-gracia").value),
      epp_requerido:epp,
    })});
    toast("✓ Configuración guardada","ok");
  }catch(e){toast("✗ Error: "+e.message,"err");}
}

// ────────────────────────────────────────────────
// PORTAL EMPLEADO
// ────────────────────────────────────────────────
function _syncEmpCamFeed(slot, camRunning){
  // Sincroniza la visibilidad del feed de cámara en el portal empleado
  if(slot===0){
    const feedWrap=document.getElementById("emp-cam1-feed");
    const offMsg=document.getElementById("emp-cam1-off");
    const feedImg=document.getElementById("emp-feed1");
    if(camRunning){
      if(!feedImg.src||feedImg.src===window.location.href){
        feedImg.src="/video/cam1?"+Date.now();
      }
      feedWrap.style.display="block";
      offMsg.style.display="none";
    }else{
      feedImg.src="";
      feedWrap.style.display="none";
      offMsg.style.display="block";
    }
  }else{
    const feedWrap=document.getElementById("emp-cam2-feed");
    const offMsg=document.getElementById("emp-cam2-off");
    const feedImg=document.getElementById("emp-feed2");
    if(camRunning){
      if(!feedImg.src||feedImg.src===window.location.href){
        feedImg.src="/video/cam2?"+Date.now();
      }
      feedWrap.style.display="block";
      offMsg.style.display="none";
    }else{
      feedImg.src="";
      feedWrap.style.display="none";
      offMsg.style.display="block";
    }
  }
}

function startEmpPolling(){
  clearInterval(empPollingInterval);
  empPollingInterval=setInterval(async()=>{
    try{
      const [dOp, dCam]=await Promise.all([
        api("/api/operario"),
        api("/api/camara/estado")
      ]);
      const ops=dOp.operarios||[];
      // Op 1
      const op1=ops.find(o=>o.slot===0);
      const el1=document.getElementById("emp-op1-status");
      if(op1){
        el1.textContent=op1.nombre.toUpperCase()+" · Entrada: "+op1.entrada;
        el1.className="emp-status active";
        _syncEmpCamFeed(0, dCam.running);
      }else{
        el1.textContent="— Ninguno registrado —";
        el1.className="emp-status none";
        _syncEmpCamFeed(0, false);
      }
      // Op 2
      const op2=ops.find(o=>o.slot===1);
      const card2=document.getElementById("emp-card-op2");
      if(op2){
        card2.style.display="block";
        const el2=document.getElementById("emp-op2-status");
        el2.textContent=op2.nombre.toUpperCase()+" · Entrada: "+op2.entrada;
        el2.className="emp-status active";
        _syncEmpCamFeed(1, dCam.running2);
      }else{
        card2.style.display="none";
        _syncEmpCamFeed(1, false);
      }
    }catch{}
  },2000);
  // Ejecutar inmediatamente al iniciar
  setTimeout(async()=>{
    try{
      const [dOp,dCam]=await Promise.all([api("/api/operario"),api("/api/camara/estado")]);
      const ops=dOp.operarios||[];
      const op1=ops.find(o=>o.slot===0);
      const el1=document.getElementById("emp-op1-status");
      if(op1){el1.textContent=op1.nombre.toUpperCase()+" · Entrada: "+op1.entrada;el1.className="emp-status active";_syncEmpCamFeed(0,dCam.running);}
      else{el1.textContent="— Ninguno registrado —";el1.className="emp-status none";_syncEmpCamFeed(0,false);}
      const op2=ops.find(o=>o.slot===1);
      const card2=document.getElementById("emp-card-op2");
      if(op2){card2.style.display="block";const el2=document.getElementById("emp-op2-status");el2.textContent=op2.nombre.toUpperCase()+" · Entrada: "+op2.entrada;el2.className="emp-status active";_syncEmpCamFeed(1,dCam.running2);}
      else{card2.style.display="none";_syncEmpCamFeed(1,false);}
    }catch{}
  },300);
}

// ────────────────────────────────────────────────
// EVIDENCIAS
// ────────────────────────────────────────────────
async function cargarEvidencias(){
  const query=(document.getElementById("ev-search")?.value||"").trim();
  const results=document.getElementById("ev-results");
  results.innerHTML='<div style="color:var(--sub);font-size:12px;text-align:center;padding:40px">Cargando evidencias...</div>';
  let data;
  try{data=await api(`/api/evidencias?q=${encodeURIComponent(query)}`);}
  catch(e){results.innerHTML=`<div style="color:var(--red);text-align:center;padding:40px">Error al cargar evidencias: ${e.message}</div>`;return;}
  if(!data.length){
    results.innerHTML=`<div style="color:var(--sub);font-size:12px;text-align:center;padding:40px">${query?`No se encontraron evidencias para: "${query}"`:"No hay evidencias registradas aún. Las evidencias se generan automáticamente cuando se detecta un evento durante el turno."}</div>`;
    return;
  }
  const grupos={};
  data.forEach(ev=>{const key=`${ev.cedula}_${ev.nombre}`;if(!grupos[key])grupos[key]={cedula:ev.cedula,nombre:ev.nombre,items:[]};grupos[key].items.push(ev);});
  const TIPO_ICON={"FATIGA":"😴","TELEFONO":"📱","EPP_FALTANTE":"⛑"};
  const TIPO_COLOR={"FATIGA":"var(--red)","TELEFONO":"var(--blue)","EPP_FALTANTE":"var(--accent)"};
  let html="";
  Object.values(grupos).forEach(grupo=>{
    html+=`<div style="background:rgba(22,22,31,0.88);border:1px solid var(--border);border-radius:10px;margin-bottom:16px;overflow:hidden;backdrop-filter:blur(8px)">
      <div style="padding:12px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between">
        <div><span style="color:var(--accent);font-weight:700">👤 ${grupo.nombre.toUpperCase()}</span><span style="color:var(--sub);font-size:11px;margin-left:10px">Cédula: ${grupo.cedula} · ${grupo.items.length} clip(s)</span></div>
      </div><div style="display:flex;flex-wrap:wrap;gap:10px;padding:14px">`;
    grupo.items.forEach(ev=>{
      const tipoKey=ev.tipo.split(":")[0];
      const icon=TIPO_ICON[tipoKey]||"🎥";
      const color=TIPO_COLOR[tipoKey]||"var(--sub)";
      const tipoLabel=ev.tipo.replace(":",(": "));
      const videoUrl=`/api/evidencias/video?carpeta=${encodeURIComponent(ev._carpeta)}&archivo=${encodeURIComponent(ev.archivo)}`;
      if(ev._existe){
        html+=`<div style="background:var(--panel);border:1px solid var(--border);border-radius:8px;width:220px;padding:12px;display:flex;flex-direction:column;gap:8px">
          <div style="color:${color};font-weight:700;font-size:12px">${icon} ${tipoLabel}</div>
          <div style="color:var(--sub);font-size:11px">${ev.ts}</div>
          <video controls style="width:100%;border-radius:6px;background:#000;max-height:130px" src="${videoUrl}">Tu navegador no soporta video.</video>
          <a href="${videoUrl}" download="${ev.archivo}" style="display:block;text-align:center;background:var(--accent);color:#fff;border-radius:6px;padding:6px 0;font-size:11px;font-weight:700;text-decoration:none;cursor:pointer">⬇ Descargar clip</a>
        </div>`;
      }else{
        html+=`<div style="background:var(--panel);border:1px solid var(--border);border-radius:8px;width:220px;padding:12px;display:flex;flex-direction:column;gap:8px">
          <div style="color:${color};font-weight:700;font-size:12px">${icon} ${tipoLabel}</div>
          <div style="color:var(--sub);font-size:11px">${ev.ts}</div>
          <div style="color:var(--red);font-size:11px;text-align:center;padding:16px 0">Archivo no encontrado</div>
        </div>`;
      }
    });
    html+='</div></div>';
  });
  results.innerHTML=html;
}

// ────────────────────────────────────────────────
// CHECK SESIÓN AL CARGAR
// ────────────────────────────────────────────────
// Auto-guardar salida si el empleado cierra la pestaña sin registrar salida
window.addEventListener("beforeunload", ()=>{
  if(ROL==="empleado"){
    // Usar sendBeacon para garantizar que el request llega aunque se cierre la tab
    navigator.sendBeacon("/api/logout");
  }
});
(async()=>{
  try{
    const d=await api("/api/session");
    if(d.autenticado){ ROL=d.rol; initApp(ROL); }
  }catch{}
})();
</script>
</body>
</html>
"""
# ── Rutas HTML ────────────────────────────────────────────────
@app.route("/")
def index():
    return INDEX_HTML


# ── Estado modelos ───────────────────────────────────────────
@app.route("/api/modelos_estado")
def api_modelos_estado():
    return jsonify({
        "listos":  modelos_listos,
        "yolo":    YOLO_OK,
        "epp":     EPP_MODELO_DISPONIBLE,
        "face":    face_mesh is not None,
    })

# ── Auth ──────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.json
    user = (data.get("usuario") or "").strip().lower()
    pwd  = (data.get("password") or "").strip()
    if user in USUARIOS and USUARIOS[user]["password"] == pwd:
        session["usuario"] = user
        session["rol"]     = USUARIOS[user]["rol"]
        return jsonify({"ok": True, "rol": USUARIOS[user]["rol"],
                        "modelos_listos": modelos_listos})
    return jsonify({"ok": False, "error": "Usuario o contraseña incorrectos"})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    # Registrar salida automática de cualquier operario activo de esta sesión
    for slot in range(2):
        op = operarios[slot]
        if op:
            hora_salida = datetime.now()
            dur = 0.0
            if horas_entrada[slot]:
                dur = round((hora_salida - horas_entrada[slot]).total_seconds() / 60, 1)
            with open(CSV_REGISTROS, "a", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow([
                    op["cedula"], op["nombre"], op["cargo"],
                    horas_entrada[slot].strftime("%Y-%m-%d %H:%M:%S") if horas_entrada[slot] else "",
                    hora_salida.strftime("%Y-%m-%d %H:%M:%S"),
                    dur])
            guardar_evento("SALIDA_AUTOMATICA", cedula=op["cedula"])
            # Detener cámara del slot
            cam = cam1 if slot == 0 else cam2
            cam["running"] = False
            if cam.get("cap"):
                cam["cap"].release()
                cam["cap"] = None
            cam["frame_jpg"] = None
            operarios[slot] = None
            horas_entrada[slot] = None
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/session")
def api_session():
    if "usuario" in session:
        return jsonify({"autenticado": True,
                        "usuario": session["usuario"],
                        "rol": session["rol"]})
    return jsonify({"autenticado": False})


# ── Operarios ─────────────────────────────────────────────────
@app.route("/api/entrada", methods=["POST"])
@login_required()
def api_entrada():
    d = request.json
    cedula = (d.get("cedula") or "").strip()
    nombre = (d.get("nombre") or "").strip()
    cargo  = (d.get("cargo")  or "").strip()
    if not cedula or not nombre or not cargo:
        return jsonify({"ok": False, "error": "Campos vacíos"})
    # Verificar que la cédula no esté ya registrada
    for op in operarios:
        if op and op["cedula"] == cedula:
            return jsonify({"ok": False, "error": "Ese operario ya tiene una entrada activa"})
    # Asignar al slot libre
    slot_asignado = None
    if operarios[0] is None:
        slot_asignado = 0
    elif operarios[1] is None:
        slot_asignado = 1
    else:
        return jsonify({"ok": False, "error": "Ya hay 2 operarios activos. Registre una salida primero."})
    datos = {"cedula": cedula, "nombre": nombre, "cargo": cargo}
    registrar_entrada_fn(datos, slot_asignado)
    return jsonify({"ok": True, "slot": slot_asignado})

@app.route("/api/salida", methods=["POST"])
@login_required()
def api_salida():
    d = request.json
    cedula = (d.get("cedula") or "").strip()
    if not cedula:
        return jsonify({"ok": False, "error": "Ingresa la cédula"})
    slot = registrar_salida_fn(cedula)
    if slot is None:
        activos = [op["cedula"] for op in operarios if op]
        return jsonify({"ok": False,
                        "error": f"Cédula no encontrada. Activos: {activos}"})
    # Detener la cámara correspondiente al slot
    cam = cam1 if slot == 0 else cam2
    cam["running"] = False
    time.sleep(0.2)
    if cam.get("cap"):
        cam["cap"].release()
        cam["cap"] = None
    cam["frame_jpg"] = None
    return jsonify({"ok": True, "slot": slot})

@app.route("/api/operario")
@login_required()
def api_operario():
    resultado = []
    for slot in range(2):
        op = operarios[slot]
        if op:
            ts = horas_entrada[slot].strftime("%H:%M:%S") if horas_entrada[slot] else "--"
            resultado.append({"slot": slot, "activo": True,
                               "nombre": op["nombre"], "cargo": op["cargo"],
                               "cedula": op["cedula"], "entrada": ts})
    # compatibilidad: si solo hay 1, retorna también activo/nombre/cargo
    if resultado:
        r = resultado[0]
        return jsonify({"activo": True,
                        "nombre": r["nombre"], "cargo": r["cargo"],
                        "cedula": r["cedula"], "entrada": r["entrada"],
                        "operarios": resultado})
    return jsonify({"activo": False, "operarios": []})


# ── Cámaras ───────────────────────────────────────────────────
def _iniciar_cam(numero):
    """Inicia la cámara 1 (numero=1) o 2 (numero=2). Retorna dict con ok/error."""
    if numero == 1:
        if cam1["running"]:
            return {"ok": True, "msg": "Cámara 1 ya estaba activa"}
        cap = cv2.VideoCapture(CAM_OPERADOR_1)
        if not cap.isOpened():
            return {"ok": False, "error": f"No se pudo abrir cámara {CAM_OPERADOR_1}"}
        cam1["cap"]     = cap
        cam1["running"] = True
        cam1["thread"]  = threading.Thread(target=_loop_cam1, daemon=True)
        cam1["thread"].start()
        return {"ok": True, "msg": "Cámara 1 iniciada"}
    else:
        if cam2["running"]:
            return {"ok": True, "msg": "Cámara 2 ya estaba activa"}
        cap2 = cv2.VideoCapture(CAM_OPERADOR_2)
        if not cap2.isOpened():
            return {"ok": False, "error": f"No se pudo abrir cámara {CAM_OPERADOR_2}"}
        cam2["cap"]     = cap2
        cam2["running"] = True
        cam2["thread"]  = threading.Thread(target=_loop_cam2, daemon=True)
        cam2["thread"].start()
        return {"ok": True, "msg": "Cámara 2 iniciada"}

@app.route("/api/camara/iniciar", methods=["POST"])
@login_required()
def api_cam_iniciar():
    """Inicia SOLO la cámara 1 (para compatibilidad con admin que la pide sola)."""
    r1 = _iniciar_cam(1)
    return jsonify(r1)

@app.route("/api/camara/iniciar/1", methods=["POST"])
@login_required()
def api_cam_iniciar_1():
    return jsonify(_iniciar_cam(1))

@app.route("/api/camara/iniciar/2", methods=["POST"])
@login_required()
def api_cam_iniciar_2():
    return jsonify(_iniciar_cam(2))

@app.route("/api/camara/detener", methods=["POST"])
@login_required()
def api_cam_detener():
    cam1["running"] = False
    cam2["running"] = False
    time.sleep(0.3)
    for cam in (cam1, cam2):
        if cam.get("cap"):
            cam["cap"].release()
            cam["cap"] = None
        cam["frame_jpg"] = None
    return jsonify({"ok": True})

@app.route("/api/camara/detener/1", methods=["POST"])
@login_required()
def api_cam_detener_1():
    cam1["running"] = False
    time.sleep(0.2)
    if cam1.get("cap"):
        cam1["cap"].release()
        cam1["cap"] = None
    cam1["frame_jpg"] = None
    return jsonify({"ok": True})

@app.route("/api/camara/detener/2", methods=["POST"])
@login_required()
def api_cam_detener_2():
    cam2["running"] = False
    time.sleep(0.2)
    if cam2.get("cap"):
        cam2["cap"].release()
        cam2["cap"] = None
    cam2["frame_jpg"] = None
    return jsonify({"ok": True})

@app.route("/video/cam1")
def video_cam1():
    return Response(gen_cam1(),
                    mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/video/cam2")
def video_cam2():
    return Response(gen_cam2(),
                    mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/api/camara/estado")
@login_required()
def api_cam_estado():
    c1 = cam1
    c2 = cam2
    alertas = []
    # Alertas cam1
    if "FATIGA" in c1["ojos"]:  alertas.append("⚠ [CAM1] FATIGA DETECTADA")
    if c1["telefono"]:           alertas.append("⚠ [CAM1] USO DE TELÉFONO")
    if c1.get("epp_falta"):      alertas.append(f"⚠ [CAM1] EPP FALTANTE: {', '.join(c1['epp_falta'])}")
    # Alertas cam2
    if c2["running"]:
        if "FATIGA" in c2.get("ojos",""):  alertas.append("⚠ [CAM2] FATIGA DETECTADA")
        if c2.get("telefono"):              alertas.append("⚠ [CAM2] USO DE TELÉFONO")
        if c2.get("epp_falta"):            alertas.append(f"⚠ [CAM2] EPP FALTANTE: {', '.join(c2['epp_falta'])}")
    return jsonify({
        "running":        c1["running"],
        "running2":       c2["running"],
        # cam1
        "ojos":           c1["ojos"],
        "ear":            round(c1["ear"], 2),
        "persona":        c1["persona"],
        "telefono":       c1["telefono"],
        "epp_ok":         list(c1["epp_ok"]),
        "epp_falta":      list(c1.get("epp_falta", set())),
        # cam2
        "ojos2":          c2.get("ojos", "SIN DETECCION"),
        "ear2":           round(c2.get("ear", 0.0), 2),
        "persona2":       c2.get("persona", False),
        "telefono2":      c2.get("telefono", False),
        "epp_ok2":        list(c2.get("epp_ok", set())),
        "epp_falta2":     list(c2.get("epp_falta", set())),
        "alertas":        alertas,
        "modelos_listos": modelos_listos,
    })


# ── Registros ─────────────────────────────────────────────────
@app.route("/api/registros")
@login_required()
def api_registros():
    return jsonify(cargar_registros())

@app.route("/api/eventos")
@login_required()
def api_eventos():
    return jsonify(cargar_eventos())

@app.route("/api/exportar/registros_excel")
@login_required(rol="administrador")
def api_exportar_registros_excel():
    path, err = _exportar_excel_registros()
    if err:
        return jsonify({"error": err}), 500
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(path, as_attachment=True,
                     download_name=f"SecureWork_Registros_{ts}.xlsx")

@app.route("/api/exportar/eventos_excel")
@login_required(rol="administrador")
def api_exportar_eventos_excel():
    path, err = _exportar_excel_eventos()
    if err:
        return jsonify({"error": err}), 500
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(path, as_attachment=True,
                     download_name=f"SecureWork_Eventos_{ts}.xlsx")



# ── Estadísticas ──────────────────────────────────────────────
@app.route("/api/estadisticas")
@login_required(rol="administrador")
def api_estadisticas():
    periodo     = request.args.get("periodo", "Todo")
    query_emp   = request.args.get("empleado", "").strip().lower()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    eventos_r   = cargar_eventos()
    registros_r = cargar_registros()

    if query_emp:
        cedula_emp = None
        nombre_emp = None
        for r in registros_r:
            if query_emp in r.get("Nombre","").lower() or query_emp in r.get("Cedula","").lower():
                cedula_emp = r.get("Cedula")
                nombre_emp = r.get("Nombre")
                break
        if not cedula_emp:
            return jsonify({"error": "Empleado no encontrado"}), 404
        eventos_r  = [e for e in eventos_r if e.get("cedula") == cedula_emp]
        registros_r= [r for r in registros_r if r.get("Cedula") == cedula_emp]

    eventos, registros = _filtrar_por_periodo(
        eventos_r, registros_r, periodo,
        fecha_desde=fecha_desde or None,
        fecha_hasta=fecha_hasta or None)

    conteo      = defaultdict(int)
    epp_detalle = defaultdict(int)
    for ev in eventos:
        tipo  = ev["tipo"]
        clave = tipo.split(":")[0]
        conteo[clave] += 1
        if tipo.startswith("EPP_FALTANTE:"):
            for item in tipo.split(":")[1].split(","):
                epp_detalle[item.strip()] += 1

    duraciones = []
    for r in registros:
        try:
            d = float(r.get("Duracion_min", 0))
            if d > 0: duraciones.append(d)
        except (ValueError, TypeError):
            pass

    if fecha_desde or fecha_hasta:
        label_periodo = f"{fecha_desde or '...'} → {fecha_hasta or '...'}"
    else:
        label_periodo = periodo

    total_min    = sum(duraciones)
    horas_totales = round(total_min / 60, 1)

    # Nombre del empleado si aplica
    nombre_emp_resp = None
    cedula_emp_resp = None
    cargo_emp_resp  = None
    if query_emp and registros:
        nombre_emp_resp = registros[0].get("Nombre", "")
        cedula_emp_resp = registros[0].get("Cedula", "")
        cargo_emp_resp  = registros[0].get("Cargo", "")

    return jsonify({
        "conteo":        dict(conteo),
        "epp_detalle":   dict(epp_detalle),
        "n_turnos":      len(registros),
        "duraciones":    duraciones,
        "horas_totales": horas_totales,
        "periodo":       label_periodo,
        "fecha_desde":   fecha_desde,
        "fecha_hasta":   fecha_hasta,
        "nombre_emp":    nombre_emp_resp,
        "cedula_emp":    cedula_emp_resp,
        "cargo_emp":     cargo_emp_resp,
        "registros_emp": registros if query_emp else [],
    })

@app.route("/api/empleados")
@login_required(rol="administrador")
def api_empleados():
    registros = cargar_registros()
    vistos = set()
    empleados = []
    for r in registros:
        key = r.get("Cedula","")
        if key and key not in vistos:
            vistos.add(key)
            empleados.append({"cedula": key, "nombre": r.get("Nombre",""),
                              "cargo": r.get("Cargo","")})
    return jsonify(empleados)


@app.route("/api/evidencias")
@login_required(rol="administrador")
def api_evidencias():
    query = request.args.get("q", "").strip().lower()
    base  = os.path.join(os.path.expanduser("~"), "Desktop",
                         "SecureWork_Descargas", "evidencias")
    resultado = []
    if not os.path.isdir(base):
        return jsonify([])
    for emp_dir in os.listdir(base):
        idx_path = os.path.join(base, emp_dir, "evidencias.json")
        if not os.path.exists(idx_path):
            continue
        try:
            with open(idx_path, "r", encoding="utf-8") as fh:
                items = json.load(fh)
            for item in items:
                item["_carpeta"] = os.path.join(base, emp_dir)
                item["_existe"]  = os.path.isfile(
                    os.path.join(base, emp_dir, item.get("archivo", "")))
                resultado.append(item)
        except Exception:
            pass
    resultado.sort(key=lambda x: x.get("ts", ""), reverse=True)
    if query:
        resultado = [e for e in resultado
                     if query in e.get("nombre", "").lower()
                     or query in e.get("cedula", "").lower()]
    return jsonify(resultado)


@app.route("/api/evidencias/video")
@login_required(rol="administrador")
def api_evidencias_video():
    """Sirve un clip de video directamente al navegador."""
    carpeta  = request.args.get("carpeta", "")
    archivo  = request.args.get("archivo", "")
    if not carpeta or not archivo:
        return jsonify({"error": "Parámetros faltantes"}), 400
    ruta = os.path.join(carpeta, archivo)
    if not os.path.isfile(ruta):
        return jsonify({"error": "Archivo no encontrado"}), 404
    return send_file(ruta, mimetype="video/x-msvideo", as_attachment=False)


# ── Análisis ISO 45001 ────────────────────────────────────────
def _generar_analisis_iso45001(conteo, epp_detalle, registros, nombre_emp=None):
    fatiga   = conteo.get("FATIGA", 0)
    telefono = conteo.get("TELEFONO", 0)
    casco    = epp_detalle.get("CASCO", 0)
    chaleco  = epp_detalle.get("CHALECO", 0)
    n_turnos = len(registros)

    lineas = []
    lineas.append("ANALISIS DE CUMPLIMIENTO — NORMA ISO 45001:2018")
    lineas.append("Sistema de Gestion de Seguridad y Salud en el Trabajo")
    lineas.append("=" * 60)
    if nombre_emp:
        lineas.append(f"Empleado evaluado: {nombre_emp}")
    lineas.append(f"Turnos analizados: {n_turnos}")
    lineas.append("")

    # 1. Incumplimientos
    lineas.append("1. INCUMPLIMIENTOS DETECTADOS")
    lineas.append("-" * 40)
    incumplimientos = []
    if casco > 0:
        incumplimientos.append(
            f"  [ISO 45001 — Clausula 8.1.2 / 8.4]\n"
            f"  Ausencia de casco de seguridad: {casco} evento(s).\n"
            f"  Incumple el control operacional de EPP obligatorio.")
    if chaleco > 0:
        incumplimientos.append(
            f"  [ISO 45001 — Clausula 8.1.2 / 8.4]\n"
            f"  Ausencia de chaleco de seguridad: {chaleco} evento(s).\n"
            f"  Incumple el uso de equipo de proteccion personal requerido.")
    if fatiga > 0:
        incumplimientos.append(
            f"  [ISO 45001 — Clausula 6.1.2 / 8.1]\n"
            f"  Eventos de fatiga detectados: {fatiga}.\n"
            f"  Indica posible incumplimiento de gestion de la jornada laboral\n"
            f"  y vigilancia del estado fisico del trabajador.")
    if telefono > 0:
        incumplimientos.append(
            f"  [ISO 45001 — Clausula 8.1 / A.8.1]\n"
            f"  Uso de dispositivo movil durante el turno: {telefono} evento(s).\n"
            f"  Incumple los controles de distraccion y uso aceptable\n"
            f"  de equipos durante operaciones de riesgo.")
    if not incumplimientos:
        lineas.append("  No se detectaron incumplimientos en el periodo analizado.")
    else:
        for item in incumplimientos:
            lineas.append(item)
            lineas.append("")

    # 2. Riesgos
    lineas.append("")
    lineas.append("2. RIESGOS, PELIGROS Y POSIBLES INCIDENTES")
    lineas.append("-" * 40)
    riesgos = []
    if casco > 0:
        riesgos.append(
            "  RIESGO FISICO — Traumatismo craneoencefalico:\n"
            "  La ausencia de casco expone al trabajador a lesiones graves\n"
            "  en la cabeza por caida de objetos, golpes o proyecciones.\n"
            "  Clasificacion: RIESGO ALTO (puede generar accidente grave o mortal).")
    if chaleco > 0:
        riesgos.append(
            "  RIESGO DE VISIBILIDAD — Atropello o colision:\n"
            "  Sin chaleco reflectivo, el trabajador no es facilmente visible\n"
            "  en entornos de baja iluminacion o con maquinaria en movimiento.\n"
            "  Clasificacion: RIESGO ALTO.")
    if fatiga > 0:
        riesgos.append(
            "  RIESGO ERGONOMICO / PSICOSOCIAL — Accidente por fatiga:\n"
            "  La fatiga reduce el tiempo de reaccion y la capacidad de\n"
            "  tomar decisiones, aumentando el riesgo de accidentes operativos.\n"
            "  Clasificacion: RIESGO MUY ALTO (factor causal en el 20% de\n"
            "  accidentes laborales segun la OIT).")
    if telefono > 0:
        riesgos.append(
            "  RIESGO CONDUCTUAL — Distraccion y error humano:\n"
            "  El uso del movil genera distraccion cognitiva que puede\n"
            "  derivar en contacto con maquinaria, caidas o lesiones.\n"
            "  Clasificacion: RIESGO ALTO.")
    if not riesgos:
        lineas.append("  No se identificaron riesgos criticos en el periodo analizado.")
    else:
        for r in riesgos:
            lineas.append(r)
            lineas.append("")

    # 3. Mejoras
    lineas.append("")
    lineas.append("3. MEJORAS Y ACCIONES CORRECTIVAS RECOMENDADAS")
    lineas.append("-" * 40)
    mejoras = []
    if casco > 0 or chaleco > 0:
        mejoras.append(
            "  a) Reforzar el procedimiento de verificacion de EPP al ingreso\n"
            "     al area de trabajo (checklist de EPP pre-turno).\n"
            "  b) Implementar puntos de control fisico en accesos a zonas\n"
            "     de riesgo con verificacion visual obligatoria.\n"
            "  c) Capacitar al personal en la importancia del EPP segun\n"
            "     ISO 45001 Clausula 7.2 (Competencia y toma de conciencia).")
    if fatiga > 0:
        mejoras.append(
            "  d) Revisar la distribucion de turnos y periodos de descanso\n"
            "     en cumplimiento del Decreto 2663 de 1950 (Codigo Sustantivo\n"
            "     del Trabajo) y las recomendaciones de la OIT.\n"
            "  e) Implementar pausas activas obligatorias cada 2 horas.\n"
            "  f) Evaluar la carga de trabajo y realizar seguimiento medico\n"
            "     al empleado con eventos de fatiga recurrentes.")
    if telefono > 0:
        mejoras.append(
            "  g) Establecer politica de zona sin movil en areas de\n"
            "     operacion critica, con senalizacion visible.\n"
            "  h) Incluir en el reglamento interno la prohibicion del uso\n"
            "     de dispositivos moviles durante la operacion.")
    if not mejoras:
        lineas.append(
            "  El desempeno es satisfactorio. Se recomienda mantener\n"
            "  las practicas actuales y realizar auditorias periodicas.")
    else:
        for m in mejoras:
            lineas.append(m)
            lineas.append("")

    # 4. Conclusión
    lineas.append("")
    lineas.append("4. CONCLUSION GENERAL")
    lineas.append("-" * 40)
    total = fatiga + casco + chaleco + telefono
    if total == 0:
        nivel = "SATISFACTORIO"
        concl = ("No se registraron incidencias en el periodo. El nivel de cumplimiento\n"
                 "con la ISO 45001 es satisfactorio. Se recomienda mantener los controles\n"
                 "actuales y realizar auditorias preventivas periodicamente.")
    elif total <= 3:
        nivel = "ACEPTABLE CON OBSERVACIONES"
        concl = ("Se detectaron incidencias menores. Se recomienda atender las\n"
                 "acciones correctivas propuestas en el punto 3 de este informe\n"
                 "para mantener el cumplimiento de la ISO 45001.")
    elif total <= 10:
        nivel = "DEFICIENTE — REQUIERE INTERVENCION"
        concl = ("El nivel de incidencias es preocupante. Se requiere implementar\n"
                 "de forma inmediata las acciones correctivas y notificar al\n"
                 "responsable del SG-SST para seguimiento y plan de mejora.")
    else:
        nivel = "CRITICO — INTERVENCION URGENTE"
        concl = ("El numero de incidencias es critico. Se recomienda suspender\n"
                 "temporalmente las actividades de riesgo, realizar una auditoria\n"
                 "interna urgente y aplicar el protocolo de mejora continua\n"
                 "estipulado en ISO 45001 Clausula 10.3.")

    lineas.append(f"  Nivel de cumplimiento: {nivel}")
    lineas.append(f"  {concl}")
    lineas.append("")
    lineas.append("  Este informe fue generado automaticamente por SecureWork.")
    lineas.append(f"  Fecha de generacion: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    return "\n".join(lineas)


# ── PDF ───────────────────────────────────────────────────────
@app.route("/api/exportar/pdf")
@login_required(rol="administrador")
def api_exportar_pdf():
    periodo     = request.args.get("periodo", "Todo")
    query_emp   = request.args.get("empleado", "").strip().lower()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()

    eventos_r   = cargar_eventos()
    registros_r = cargar_registros()
    nombre_emp  = None

    if query_emp:
        for r in registros_r:
            if query_emp in r.get("Nombre","").lower() or query_emp in r.get("Cedula","").lower():
                nombre_emp = r.get("Nombre")
                cedula_emp = r.get("Cedula")
                break
        if not nombre_emp:
            return jsonify({"error": "Empleado no encontrado"}), 404
        eventos_r  = [e for e in eventos_r if e.get("cedula") == cedula_emp]
        registros_r= [r for r in registros_r if r.get("Cedula") == cedula_emp]

    eventos, registros = _filtrar_por_periodo(
        eventos_r, registros_r, periodo,
        fecha_desde=fecha_desde or None,
        fecha_hasta=fecha_hasta or None)

    if fecha_desde or fecha_hasta:
        label_periodo = f"{fecha_desde or '...'} → {fecha_hasta or '...'}"
    else:
        label_periodo = periodo

    conteo      = defaultdict(int)
    epp_detalle = defaultdict(int)
    for ev in eventos:
        tipo  = ev["tipo"]
        clave = tipo.split(":")[0]
        conteo[clave] += 1
        if tipo.startswith("EPP_FALTANTE:"):
            for item in tipo.split(":")[1].split(","):
                epp_detalle[item.strip()] += 1

    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Image as RLImage,
                                        PageBreak, HRFlowable, Table, TableStyle)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    except ImportError:
        return jsonify({"error": "reportlab no instalado. Ejecuta: pip install reportlab"}), 500

    # ── Generar gráficas matplotlib ───────────────────────────
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig_paths = []

    C_BG_PLT   = "#141414"
    C_RED_PLT  = "#ff3d3d"
    C_ACC_PLT  = "#ff6b00"
    C_BLU_PLT  = "#2196f3"
    C_GRN_PLT  = "#00c853"
    C_SUB_PLT  = "#888888"
    C_YEL_PLT  = "#ffc107"
    C_TXT_PLT  = "#e8e8e8"
    C_BOR_PLT  = "#2a2a2a"

    # Gráfico 1 — Barras de eventos por tipo
    if conteo:
        labels = list(conteo.keys())
        values = list(conteo.values())
        bar_colors = []
        for lbl in labels:
            if "FATIGA"   in lbl: bar_colors.append(C_RED_PLT)
            elif "EPP"    in lbl: bar_colors.append(C_ACC_PLT)
            elif "TELEFONO" in lbl: bar_colors.append(C_BLU_PLT)
            elif "ENTRADA"  in lbl: bar_colors.append(C_GRN_PLT)
            elif "SALIDA"   in lbl: bar_colors.append(C_SUB_PLT)
            else:                   bar_colors.append(C_SUB_PLT)
        fig1, ax1 = plt.subplots(figsize=(9, 3.5), facecolor="white")
        ax1.set_facecolor("#f9f9f9")
        bars = ax1.bar(labels, values, color=bar_colors, edgecolor="#cccccc")
        ax1.set_title(f"Eventos Generales — Periodo: {label_periodo}", fontsize=10)
        ax1.tick_params(labelsize=8)
        for bar, val in zip(bars, values):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.05,
                     str(val), ha="center", va="bottom", fontsize=8)
        plt.xticks(rotation=20, ha="right")
        plt.tight_layout()
        tmp1 = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        plt.savefig(tmp1.name, dpi=120, bbox_inches="tight")
        plt.close(fig1)
        fig_paths.append(tmp1.name)

    # Gráfico 2 — Pie EPP faltante
    if epp_detalle:
        epp_labels = list(epp_detalle.keys())
        epp_values = list(epp_detalle.values())
        epp_colors = [C_ACC_PLT if "CASCO" in l else C_YEL_PLT for l in epp_labels]
        fig2, ax2 = plt.subplots(figsize=(5, 3.5), facecolor="white")
        ax2.set_facecolor("white")
        ax2.pie(epp_values, labels=epp_labels, colors=epp_colors,
                autopct="%1.0f%%", startangle=90)
        ax2.set_title("EPP Faltante por Tipo", fontsize=10)
        plt.tight_layout()
        tmp2 = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        plt.savefig(tmp2.name, dpi=120, bbox_inches="tight")
        plt.close(fig2)
        fig_paths.append(tmp2.name)

    # Gráfico 4 — Resumen de Incidencias (barras horizontales, estilo dashboard)
    fatiga_tot   = conteo.get("FATIGA",   0)
    telefono_tot = conteo.get("TELEFONO", 0)
    epp_tot      = sum(epp_detalle.values())
    inc_labels = ["EPP Faltante", "Teléfono", "Fatiga"]
    inc_values = [epp_tot, telefono_tot, fatiga_tot]
    inc_colors = ["#ff6b00", "#444444", "#ff3d5a"]
    fig4, ax4 = plt.subplots(figsize=(9, 2.8), facecolor="white")
    ax4.set_facecolor("#f9f9f9")
    bars4 = ax4.barh(inc_labels, inc_values, color=inc_colors, height=0.55, edgecolor="none")
    ax4.set_title("⚡ RESUMEN DE INCIDENCIAS", fontsize=9, color="#333333",
                  loc="left", pad=10, fontweight="bold")
    ax4.tick_params(colors="#333333", labelsize=8)
    ax4.spines[:].set_visible(False)
    ax4.xaxis.set_tick_params(color="#cccccc")
    ax4.set_xticks(range(0, max(inc_values or [1]) + 2))
    ax4.tick_params(axis="x", colors="#aaaaaa")
    ax4.tick_params(axis="y", colors="#333333")
    ax4.set_facecolor("#f9f9f9")
    fig4.patch.set_facecolor("white")
    plt.tight_layout(pad=1.2)
    tmp4 = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    plt.savefig(tmp4.name, dpi=140, bbox_inches="tight", facecolor="white")
    plt.close(fig4)
    fig_paths.insert(0, tmp4.name)   # Va primero en el PDF

    # Gráfico 3 — Duración de turnos (solo por empleado)
    if query_emp:
        duraciones_pdf = []
        for r in registros:
            try:
                d = float(r.get("Duracion_min", 0))
                if d > 0:
                    duraciones_pdf.append(d)
            except (ValueError, TypeError):
                pass
        if duraciones_pdf:
            etiquetas = [f"T{i+1}" for i in range(len(duraciones_pdf))]
            fig3, ax3 = plt.subplots(figsize=(9, 3), facecolor="white")
            ax3.set_facecolor("#f9f9f9")
            ax3.bar(etiquetas, duraciones_pdf, color=C_BLU_PLT, edgecolor="#cccccc")
            ax3.set_title(f"Duración de Turnos (min) — {nombre_emp}", fontsize=10)
            ax3.set_ylabel("Minutos", fontsize=8)
            ax3.tick_params(labelsize=7)
            plt.tight_layout()
            tmp3 = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            plt.savefig(tmp3.name, dpi=120, bbox_inches="tight")
            plt.close(fig3)
            fig_paths.append(tmp3.name)

    tmp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    doc = SimpleDocTemplate(tmp_pdf.name, pagesize=letter,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2.5*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    st_title = ParagraphStyle("T", parent=styles["Title"], fontSize=18,
                              textColor=colors.HexColor("#FF6B00"),
                              fontName="Helvetica-Bold", alignment=TA_CENTER)
    st_sub   = ParagraphStyle("S", parent=styles["Normal"], fontSize=10,
                              textColor=colors.HexColor("#888888"), alignment=TA_CENTER)
    st_h2    = ParagraphStyle("H", parent=styles["Heading2"], fontSize=12,
                              textColor=colors.HexColor("#FF6B00"),
                              fontName="Helvetica-Bold", spaceBefore=14)
    st_body  = ParagraphStyle("B", parent=styles["Normal"], fontSize=9,
                              textColor=colors.HexColor("#1a1a1a"),
                              leading=14, alignment=TA_JUSTIFY)

    story = [Spacer(1, 1*cm)]
    story.append(Paragraph("SecureWork", st_title))
    story.append(Paragraph("Sistema de Gestión de Seguridad y Salud Ocupacional", st_sub))
    story.append(HRFlowable(width="100%", thickness=2,
                             color=colors.HexColor("#FF6B00"), spaceAfter=8))
    titulo = f"Informe {'— ' + nombre_emp if nombre_emp else 'General'} — {label_periodo}"
    story.append(Paragraph(titulo, ParagraphStyle("T2", parent=styles["Heading1"],
                                                   fontSize=14,
                                                   textColor=colors.HexColor("#222222"),
                                                   alignment=TA_CENTER)))
    story.append(Paragraph(
        f"Periodo: {label_periodo}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        st_sub))
    story.append(Spacer(1, 0.6*cm))

    story.append(Paragraph("Resumen de Indicadores", st_h2))
    kpi_data = [
        ["Indicador","Valor","Estado"],
        ["Turnos registrados", str(len(registros)), "INFO"],
        ["Eventos de fatiga", str(conteo.get("FATIGA",0)),
         "CRITICO" if conteo.get("FATIGA",0)>5 else ("ALERTA" if conteo.get("FATIGA",0)>0 else "OK")],
        ["Casco faltante",  str(epp_detalle.get("CASCO",0)),
         "CRITICO" if epp_detalle.get("CASCO",0)>3 else ("ALERTA" if epp_detalle.get("CASCO",0)>0 else "OK")],
        ["Chaleco faltante",str(epp_detalle.get("CHALECO",0)),
         "CRITICO" if epp_detalle.get("CHALECO",0)>3 else ("ALERTA" if epp_detalle.get("CHALECO",0)>0 else "OK")],
        ["Uso de teléfono", str(conteo.get("TELEFONO",0)),
         "ALERTA" if conteo.get("TELEFONO",0)>0 else "OK"],
    ]
    color_map = {"OK": colors.HexColor("#00C853"), "ALERTA": colors.HexColor("#FFC107"),
                 "CRITICO": colors.HexColor("#FF3D3D"), "INFO": colors.HexColor("#2196F3")}
    tbl = Table(kpi_data, colWidths=[9*cm, 3*cm, 4.5*cm])
    tbl_style = TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#FF6B00")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("ALIGN",      (0,0),(-1,-1),"CENTER"),
        ("VALIGN",     (0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#f9f9f9"),colors.HexColor("#efefef")]),
        ("GRID",       (0,0),(-1,-1), 0.5, colors.HexColor("#cccccc")),
        ("FONTSIZE",   (0,1),(-1,-1), 9),
        ("TOPPADDING", (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
    ])
    for i, row in enumerate(kpi_data[1:], 1):
        col = color_map.get(row[2], colors.grey)
        tbl_style.add("TEXTCOLOR",(2,i),(2,i), col)
        tbl_style.add("FONTNAME", (2,i),(2,i), "Helvetica-Bold")
    tbl.setStyle(tbl_style)
    story.append(tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Gráficas matplotlib ──────────────────────────────────
    if fig_paths:
        story.append(Paragraph("Gráficos del Periodo", st_h2))
        for fp in fig_paths:
            if fp and os.path.exists(fp):
                try:
                    story.append(RLImage(fp, width=16*cm, height=7*cm))
                    story.append(Spacer(1, 0.3*cm))
                except Exception:
                    pass

    story.append(PageBreak())
    story.append(Paragraph("Análisis de Cumplimiento ISO 45001:2018", st_h2))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#FF6B00"), spaceAfter=8))

    st_mono = ParagraphStyle("M", parent=styles["Code"], fontSize=8,
                             textColor=colors.HexColor("#333333"),
                             backColor=colors.HexColor("#f5f5f5"),
                             spaceAfter=3, leading=13, leftIndent=10)
    st_bold_red = ParagraphStyle("BR", parent=styles["Normal"], fontSize=9,
                                 textColor=colors.HexColor("#cc3300"),
                                 spaceAfter=2, leading=13,
                                 fontName="Helvetica-Bold")

    analisis_txt = _generar_analisis_iso45001(
        dict(conteo), dict(epp_detalle), registros, nombre_emp=nombre_emp)

    for linea in analisis_txt.split("\n"):
        ls = linea.strip()
        if not ls:
            story.append(Spacer(1, 0.15*cm))
        elif ls.startswith("=") or ls.startswith("-"):
            story.append(HRFlowable(width="100%", thickness=0.5,
                                     color=colors.HexColor("#cccccc"), spaceAfter=2))
        elif len(ls) > 1 and ls[0].isdigit() and ls[1] == ".":
            story.append(Paragraph(ls, st_h2))
        elif ls.startswith("[ISO") or ls.startswith("RIESGO") or ls.startswith("Nivel"):
            story.append(Paragraph(f"<b>{ls}</b>", st_bold_red))
        else:
            story.append(Paragraph(linea, st_mono))

    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor("#cccccc"), spaceAfter=4))
    story.append(Paragraph(
        "Documento generado por SecureWork — Proyecto Académico Ingeniería Industrial "
        "— Salud Ambiental y Ocupacional",
        ParagraphStyle("Foot", parent=styles["Normal"], fontSize=7,
                       textColor=colors.HexColor("#aaaaaa"), alignment=TA_CENTER)))

    doc.build(story)
    # Limpiar archivos temporales de gráficas
    for fp in fig_paths:
        try:
            os.remove(fp)
        except Exception:
            pass
    return send_file(tmp_pdf.name, as_attachment=True,
                     download_name=f"informe_securework_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

# ── Configuración ─────────────────────────────────────────────
@app.route("/api/config", methods=["GET"])
@login_required(rol="administrador")
def api_config_get():
    return jsonify({
        "ear_umbral":     EAR_UMBRAL,
        "tiempo_umbral":  TIEMPO_UMBRAL,
        "epp_gracia":     EPP_TIEMPO_GRACIA,
        "conf_umbral":    model_epp.overrides.get("conf", 0.25) if model_epp else 0.25,
        "epp_requerido":  list(EPP_REQUERIDO),
        "cam1_index":     CAM_OPERADOR_1,
        "cam2_index":     CAM_OPERADOR_2,
    })

@app.route("/api/config", methods=["POST"])
@login_required(rol="administrador")
def api_config_post():
    global EAR_UMBRAL, TIEMPO_UMBRAL, EPP_TIEMPO_GRACIA, EPP_REQUERIDO
    global CAM_OPERADOR_1, CAM_OPERADOR_2
    d = request.json
    EAR_UMBRAL        = float(d.get("ear_umbral",     EAR_UMBRAL))
    TIEMPO_UMBRAL     = float(d.get("tiempo_umbral",  TIEMPO_UMBRAL))
    EPP_TIEMPO_GRACIA = float(d.get("epp_gracia",     EPP_TIEMPO_GRACIA))
    CAM_OPERADOR_1    = int(d.get("cam1_index",   CAM_OPERADOR_1))
    CAM_OPERADOR_2    = int(d.get("cam2_index",   CAM_OPERADOR_2))
    if model_epp:
        model_epp.overrides["conf"] = float(d.get("conf_umbral", 0.25))
    epp_req = d.get("epp_requerido", [])
    EPP_REQUERIDO = set(epp_req)
    est_casco.reconf(
        int(d.get("frames_act_casco", 8)),
        int(d.get("frames_des_casco", 30))
    )
    est_chaleco.reconf(
        int(d.get("frames_act_chaleco", 8)),
        int(d.get("frames_des_chaleco", 30))
    )
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
